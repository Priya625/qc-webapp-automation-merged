import pandas as pd
import re
from typing import List ,Dict,Any, Set
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO 
from pandas.api.types import is_object_dtype, is_categorical_dtype, CategoricalDtype 
from fuzzywuzzy import fuzz
from datetime import datetime, timedelta
import numpy as np
from datetime import timedelta


# --- Constants ---
DATE_FORMAT = "%Y-%m-%d"
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


class EPLValidator:
    """
    Handles loading, validating, and processing of BSR data.
    The dependency on the Rosco file has been removed.
    
    """
    # --- AUDIENCE CHECK CLASS CONSTANTS ---
    OVERNIGHT_SHEET = "DATA"
    OVERNIGHT_AUDIENCE_COL = 'Audience'
    BSR_TARGET_COL_RAW = 'Aud Metered (000s) 3+'
    GP_FILTER_COL = 'Grand Prix'
    GP_FILTER_VALUE = '15_Dutch GP'
    
    # Canonical Column Names
    COUNTRY_COLUMN = 'Market' 
    CHANNEL_COLUMN = 'TV-Channel'
    DATE_COLUMN = 'Date'
    SESSION_COMPETITION_COLUMN = 'Competition'

    def _define_check_defaults(self):
        """Defines the default parameters. These are used if the user doesn't provide input."""
        return {
            "impute_lt_live_status": {
                "market": "INDIA",
                "keyword": "L/T",
                "required_cols": ["Combined", "Type of program", "Market"]
            },
            "consolidate_gillete_soccer": {
                "keyword": "GILLETE SOCCER",
                "max_gap_minutes": 30,
                "required_cols": ["Combined", "Date", "Start", "End", "Market", "TV-Channel"]
            }
        }

    def _merge_configs(self, defaults, user_inputs):
        """Smartly merges user inputs over the defaults."""
        merged = defaults.copy()
        for check_key, user_config in user_inputs.items():
            if check_key in merged and isinstance(user_config, dict):
                # Update specific keys (e.g. just change 'market' but keep 'keyword')
                merged[check_key].update(user_config)
            elif check_key in merged:
                merged[check_key] = user_config
        return merged

    def __init__(self, df: pd.DataFrame, bsr_path: str, obligation_path: str = None, overnight_path: str = None, macro_path: str = None, check_configs: Dict[str, Any] = None):
    # def __init__(self, bsr_path: str, obligation_path: str = None, overnight_path: str = None, macro_path: str = None):
        self.df = df        
        self.bsr_path = bsr_path
        self.df = self._load_bsr()
        # New: Store the obligation path, but don't load the full DF yet
        self.obligation_path = obligation_path
        self.full_obligation_df = None # Will store the entire obligation sheet
        # NEW: Store the overnight path
        self.overnight_path = overnight_path # <-- STORED HERE
        self.macro_path = macro_path
        # 🚨 NEW: Load and store the duplication rules DataFrame
        self.dup_rules_df = self._load_and_filter_macro_rules()

         # 🚨 CONFIG SETUP: Merge Defaults + User Inputs
        default_configs = self._define_check_defaults()
        # If check_configs is None, use empty dict
        self.config = self._merge_configs(default_configs, check_configs or {})

        # ✅ FIX: Load the DataFrame immediately using the path
        # try:
        #     self.df = pd.read_excel(self.bsr_path)
        # except Exception as e:
        #     raise ValueError(f"Failed to load BSR file at {self.bsr_path}: {e}")

        # Initialize other attributes
        self.full_obligation_df = None 
        
        # Load duplication rules (Assuming _load_and_filter_macro_rules is defined in your class)
        try:
            self.dup_rules_df = self._load_and_filter_macro_rules()
        except Exception:
            # Handle case where macro path might not be set or valid
            self.dup_rules_df = pd.DataFrame()
        
        # Dictionary to map market check keys to internal methods (to be implemented)
        self.market_check_map = {
        "impute_lt_live_status": self._impute_lt_live_status,
        "consolidate_gillete_soccer": self._consolidate_gillette_soccer_programs,
        "check_sky_showcase_live": self._check_sky_showcase_live_status,
        "standardize_uk_ire_region": self._standardize_uk_ire_region,
        "check_fixture_vs_case" : self._check_fixture_vs_case,
        "check_pan_balkans_serbia_parity" : self._check_pan_balkans_serbia_parity,
        "audit_multi_match_status" : self._audit_multi_match_status,
        "check_date_time_format_integrity" : self._check_date_time_format_integrity,
        "check_live_broadcast_uniqueness" : self._check_live_broadcast_uniqueness,
        "audit_channel_line_item_count" : self._audit_channel_line_item_count,
        "check_combined_archive_status": self._check_combined_archive_status,
        "suppress_duplicated_audience" : self._suppress_duplicated_audience,
        "harmonize_uk_ire_program_descriptions_strict" : self._harmonize_uk_ire_program_descriptions_simple,
        "check_game_of_the_day_match" : self._check_game_of_the_day_match,
        "check_non_metered_primary_market_audience" : self._check_non_metered_primary_market_audience,
        "check_legacy_mapping" : self._check_legacy_mapping,
        "check_premier_league_october_obligation" : self._check_premier_league_october_obligation,
        #"check_star_sports_3_consolidation" : self._check_star_sports_3_consolidation,
        "check_bsa_nielsen_audience_presence" : self._check_bsa_nielsen_audience_presence,
        "check_source_mediatype_validity" : self._check_source_mediatype_validity,
        "filter_short_programs": self._filter_short_programs,
        "sa_nielsen_inclusion_check": self._sa_nielsen_inclusion_check,
        "epl_live_vs_delay_validation": self._epl_live_vs_delay_validation,
        "pl_magazine_highlights_classification": self._pl_magazine_highlights_classification,
        #"relevant_only_in_the_uk": self._relevant_only_in_the_uk,
        #"dedicated_program_duration_allignments": self._dedicated_program_duration_allignments
        # Future EPL checks would be added here
    }

    def _load_and_filter_macro_rules(self):
        """Loads, filters, and standardizes the macro duplication rules file."""
        if not self.macro_path:
            return None
            
        MACRO_SHEET_NAME = "Data Core"
        MACRO_HEADER_INDEX = 1 
        SEARCH_TERM = "Formula 1"
        REQUIRED_RULE_COLS = ['Orig Market', 'Dup Market', 'Dup Channel', 'Projects'] # Include Projects for filtering

        try:
            df_macro = pd.read_excel(self.macro_path, sheet_name=MACRO_SHEET_NAME, header=MACRO_HEADER_INDEX)
            df_macro.columns = [str(c).strip() for c in df_macro.columns]

            # 1. Filter by Project (Formula 1)
            filtered_df = df_macro[
                df_macro['Projects'].astype(str).str.contains(SEARCH_TERM, case=False, na=False)
            ].copy()
            
            # 2. Select and clean required columns
            df_dup_rules = filtered_df[REQUIRED_RULE_COLS].copy()

            # Ensure key columns are clean strings (strip, upper case)
            for col in ['Orig Market', 'Dup Market', 'Dup Channel']:
                if col in df_dup_rules.columns:
                    df_dup_rules[col] = df_dup_rules[col].astype(str).str.strip().str.upper()
                    
            # We only need 'Orig Market', 'Dup Market', 'Dup Channel' for the validation check
            return df_dup_rules[['Orig Market', 'Dup Market', 'Dup Channel']].drop_duplicates()
        
        except Exception as e:
            print(f"Error loading duplication rules from macro file: {e}")
            return None

    # --- Private Loading/Parsing Methods (from old qc_checks.py) ---
    def _load_overnight_data(self):
        """ Loads the CDT/Overnight file, handling the specific header offset (Row 9)
        and standardizing columns for the Game of the Day check.
        """
        if not self.overnight_path: return None
        
        # Constants for the CDT/OVN File Structure
        OVN_HEADER_ROW = 8 # Row 9 in Excel is Index 8 in Pandas
        
        # Columns we need to extract
        # Mapping: 'Date' -> Date, 'Start Time' -> Start, 'Matchweek' -> Matchday Source, 'Programme Title' -> Desc Source
        OVN_USE_COLS = ['Date', 'Start Time', 'End Time', 'Matchweek', 'Programme Title', 'Broadcaster']
        
        try:
            # Load with specific header row
            df_ovn = pd.read_excel(self.overnight_path, sheet_name=0, header=OVN_HEADER_ROW)
            
            # Normalize headers (strip spaces) to ensure we find the right columns
            df_ovn.columns = [str(c).strip() for c in df_ovn.columns]
            
            # Filter for required columns if they exist
            existing_cols = [c for c in OVN_USE_COLS if c in df_ovn.columns]
            df_ovn = df_ovn[existing_cols].copy()
            
            # --- Standardization ---
            
            # 1. Date Parsing
            if 'Date' in df_ovn.columns:
                df_ovn['Date_Clean'] = pd.to_datetime(df_ovn['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
                
            # 2. Time Parsing (using helper)
            if 'Start Time' in df_ovn.columns:
                df_ovn['Start_Time_Clean'] = self._safe_get_time_string(df_ovn['Start Time'])
            
            # 3. Create a Unique Key for Matching: Date + Start Time
            if 'Date_Clean' in df_ovn.columns and 'Start_Time_Clean' in df_ovn.columns:
                df_ovn['Match_Key'] = df_ovn['Date_Clean'] + '|' + df_ovn['Start_Time_Clean']
                
            return df_ovn
            
        except Exception as e:
            print(f"Error loading Overnight/CDT file: {e}")
            return None

    def _update_audience_from_overnight(self) -> Dict[str, Any]:
        """
        Compares BSR audience with Max Overnight Audience, updating the BSR value if 
        the overnight audience is higher, and explicitly flagging the status of every row.
        """
        initial_rows = len(self.df)
        
        # --- CONSTANTS ---
        OVERNIGHT_AUDIENCE_COL = self.OVERNIGHT_AUDIENCE_COL
        BSR_TARGET_COL_RAW = self.BSR_TARGET_COL_RAW 
        QC_FLAG_COL = 'QC_Audience_Update_Status' # NEW Status Flag Column
        
        # Canonical Column Names
        COUNTRY_COLUMN = self.COUNTRY_COLUMN      
        CHANNEL_COLUMN = self.CHANNEL_COLUMN      
        DATE_COLUMN = self.DATE_COLUMN            
        SESSION_COMPETITION_COLUMN = self.SESSION_COMPETITION_COLUMN 
        
        FINAL_MERGE_ON_COLS = [COUNTRY_COLUMN, CHANNEL_COLUMN, DATE_COLUMN, SESSION_COMPETITION_COLUMN]
        
        # 1. Load and Prepare Overnight data (Assumed correct)
        df_overnight = self._load_overnight_data()

        if df_overnight is None or BSR_TARGET_COL_RAW not in self.df.columns:
            return {"check_key": "update_audience_from_overnight", "status": "Skipped", "action": "Audience Update", "description": "Skipped: Missing Overnight file or target BSR column.", "details": {"rows_updated": 0}}
        
        # 2. Prepare BSR for merging (Standardize keys)
        self.df[BSR_TARGET_COL_RAW] = pd.to_numeric(self.df[BSR_TARGET_COL_RAW], errors='coerce')
        
        # Apply standardization to BSR columns
        for col in [COUNTRY_COLUMN, CHANNEL_COLUMN, SESSION_COMPETITION_COLUMN]:
            if col in self.df.columns:
                self.df.loc[:, col] = self.df[col].astype(str).str.strip().str.upper()
        if DATE_COLUMN in self.df.columns:
            self.df.loc[:, DATE_COLUMN] = pd.to_datetime(self.df[DATE_COLUMN], errors='coerce')
            
        # --- 3. AGGREGATE OVERNIGHT DATA (Get max audience per key) ---
        df_overnight_max = df_overnight.groupby(FINAL_MERGE_ON_COLS, dropna=False)[OVERNIGHT_AUDIENCE_COL].max().reset_index()
        df_overnight_max = df_overnight_max.rename(columns={OVERNIGHT_AUDIENCE_COL: 'Max_Overnight_Audience'})

        # 4. MERGE AND COMPARE
        merged_df = self.df.merge(
            df_overnight_max, 
            on=FINAL_MERGE_ON_COLS, 
            how='left' 
        )
        
        # Initialize the new status column in the merged DataFrame
        merged_df[QC_FLAG_COL] = 'No Match Found' # Default state

        # Scale BSR audience to absolute numbers (multiplying by 1000)
        temp_bsr_abs = merged_df[BSR_TARGET_COL_RAW] * 1000.0

        # Mask A: Rows where a match was found (Max_Overnight_Audience is NOT NaN)
        match_found_mask = merged_df['Max_Overnight_Audience'].notna()
        
        # Mask B: Rows updated (Max_Overnight_Audience > BSR_ABS)
        update_mask = match_found_mask & \
                    (merged_df['Max_Overnight_Audience'] > temp_bsr_abs) & \
                    (merged_df[BSR_TARGET_COL_RAW].notna())

        # --- 5. Apply Status Flags ---
        
        # Status 2: OK (Match found, but BSR was already higher or equal)
        # This is the residual mask: Match found AND NOT updated.
        ok_mask = match_found_mask & (~update_mask)
        merged_df.loc[ok_mask, QC_FLAG_COL] = 'OK - BSR Value Retained'
        
        # Status 1: UPDATED (The highest priority flag)
        merged_df.loc[update_mask, QC_FLAG_COL] = 'UPDATED - Scaled from Overnight Max'

        # 6. Perform the value update
        rows_updated = update_mask.sum()
        
        if rows_updated > 0:
            updated_value_in_thousands = merged_df.loc[update_mask, 'Max_Overnight_Audience'] / 1000.0
            
            # Write the new audience value to the BSR's target column
            self.df.loc[update_mask[update_mask].index, BSR_TARGET_COL_RAW] = updated_value_in_thousands 
        
        # --- 7. Finalize (Copy new columns back to self.df) ---
        self.df[QC_FLAG_COL] = merged_df[QC_FLAG_COL]

        return {
            "check_key": "update_audience_from_overnight",
            "status": "Completed" if rows_updated == 0 else "Flagged",
            "action": "Audience Update",
            "description": f"Updated BSR audience rows by overriding {rows_updated} values with higher Max Overnight data.",
            "details": {
                "rows_updated": int(rows_updated),
                "rows_not_matched": int(ok_mask.sum()),
                "rows_skipped": int((merged_df[QC_FLAG_COL] == 'No Match Found').sum()),
                "total_rows_processed": int(initial_rows)
            }
        }

    # New Private Method to load the full obligation sheet once
    # def _load_full_obligation_data(self) -> pd.DataFrame:
    #     """
    #     Loads the F1 Obligation sheet and filters it to include ONLY the '15_Dutch GP' 
    #     event data, storing the filtered DataFrame in self.full_obligation_df.
    #     """
    #     if self.full_obligation_df is not None:
    #         return self.full_obligation_df

    #     if not self.obligation_path:
    #         return pd.DataFrame()
            
    #     TARGET_GP = '15_Dutch GP' # <-- Define the target GP here
        
    #     try:
    #         # Load the entire obligation sheet
    #         df_obl = pd.read_excel(
    #             self.obligation_path, 
    #             sheet_name="F1 - Broadcaster Obligations",
    #         )
    #         df_obl.columns = [str(c).strip() for c in df_obl.columns]
            
    #         # --- CRITICAL FILTERING STEP ---
    #         # Filter the loaded DataFrame for the specific GP
    #         df_obl_filtered = df_obl[df_obl.get('GP') == TARGET_GP].copy()

    #         print(f"Obligation data loaded and filtered for: {TARGET_GP}. Rows found: {len(df_obl_filtered)}")
            
    #         # Store and return the filtered DataFrame
    #         self.full_obligation_df = df_obl_filtered
    #         return df_obl_filtered
            
    #     except FileNotFoundError:
    #         print(f"Error: Obligation file not found at {self.obligation_path}")
    #         return pd.DataFrame()
    #     except Exception as e:
    #         print(f"Error loading/filtering obligation sheet: {e}")
    #         return pd.DataFrame()

    def _load_full_obligation_data(self):
        """
        Loads the Legacy Mapping sheet, handling the specific two-row header structure
        (Row 1 = Categories, Row 2 = Columns).
        """
        if not self.obligation_path:
            return None
            
        # Assuming the sheet name is 'Legacy Mapping' or similar. 
        # You might need to adjust this if the sheet has a specific name in the file.
        LEGACY_SHEET_NAME = "BC Log Channels list" 
        
        try:
            # Load with header=[0, 1] to create a MultiIndex columns object
            # This captures 'Original details' -> 'Market' hierarchy
            df_legacy = pd.read_excel(
                self.obligation_path, 
                sheet_name=LEGACY_SHEET_NAME, 
                header=[0, 1] 
            )
            
            # Check if the expected top-level headers exist
            if 'Original details' not in df_legacy.columns.get_level_values(0):
                print(f"Error: 'Original details' header not found in {LEGACY_SHEET_NAME}")
                return None
                
            return df_legacy
            
        except Exception as e:
            # Fallback: Try loading by index if name fails, or just print error
            print(f"Error loading Legacy Mapping sheet: {e}")
            return None

    def _detect_header_row(self, sheet_name=0):
        """
        Detects the header row index by scanning the first 200 rows 
        of the specified sheet for key column names.
        
        Args:
            sheet_name: The name or index of the Excel sheet to read. Defaults to the first sheet (0).
        """
        # Read a sample of the specified sheet
        df_sample = pd.read_excel(
            self.bsr_path, 
            sheet_name=sheet_name, 
            header=None, 
            nrows=200
        )
        
        for i, row in df_sample.iterrows():
            # Convert row to a single, space-separated, lowercase string for detection
            # Use fillna('') to handle rows that might be mostly empty
            row_str = " ".join(row.fillna('').astype(str).tolist()).lower()

            # First set of keywords (common BSR columns)
            if all(k in row_str for k in ["region", "market", "broadcaster"]):
                return i
            
            # Second set of keywords (common date/time columns)
            if "date" in row_str and ("utc" in row_str or "gmt" in row_str):
                return i
                
        raise ValueError(f"Could not detect header row in '{sheet_name}' sheet of BSR file.")

    def _load_bsr(self):
        # Define the specific sheet name based on your example
        sheet_name_to_load = "Worksheet" 

        # Detect the header row on the specified sheet
        header_row = self._detect_header_row(sheet_name=sheet_name_to_load)

        # Load the full data using the detected header row and sheet name
        df = pd.read_excel(
            self.bsr_path, 
            sheet_name=sheet_name_to_load,  # Use the specific sheet name
            header=header_row               # Use the dynamically detected header row
        )
        
        # Ensure column names are clean
        df.columns = [str(c).strip() for c in df.columns]
        return df

    # --- Methods for Market Specific Checks (Placeholder Implementation) ---
    def market_check_processor(self, checks: List[str]) -> List[Dict[str, Any]]:
        # ... (Method contents remain unchanged - assumed correct)
        status_summaries = [] 
        
        for check_key in checks:
            if check_key in self.market_check_map:
                try:
                    result = self.market_check_map[check_key]()
                    if result:
                        status_summaries.append(result)
                    print(f"Applied custom check: {check_key}")
                except Exception as e:
                    status_summaries.append({
                        "check_key": check_key,
                        "status": "Failed",
                        "action": "Error during execution",
                        "description": f"Check failed due to internal error: {str(e)}",
                        "details": {"error": str(e)}
                    })
                    print(f"Error applying check {check_key}: {e}")
            else:
                print(f"Warning: Unknown check key received: {check_key}")
                
        return status_summaries

    def _impute_lt_live_status(self) -> Dict[str, Any]:
        """
        Classifies program as Live based on 'L/T' keyword in Combined column.
        RESTRICTION: Only applies if 'Home Team' and 'Away Team' are NOT blank.
        Uses CONFIGURABLE market and keyword.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Recommended_Program_Type' 
        
        # 🟢 READ FROM CONFIG
        chk_config = self.config.get("impute_lt_live_status", {})
        TARGET_MARKET = chk_config.get("market", "INDIA")
        KEYWORD = chk_config.get("keyword", "L/T")
        
        # FIX: Check for the specific spelling variant in your file
        TYPE_COL_NAME = 'Type of program'
        if 'Type of programme' in self.df.columns:
            TYPE_COL_NAME = 'Type of programme'
            
        REQUIRED_COLS = ['Combined', TYPE_COL_NAME, 'Market', 'Home Team', 'Away Team'] # Added Team cols

        self.df[FLAG_COLUMN] = 'Current Status OK'
        
        if not all(col in self.df.columns for col in REQUIRED_COLS):
             missing = [c for c in REQUIRED_COLS if c not in self.df.columns]
             return {
                 "check_key": "impute_lt_live_status", 
                 "status": "Skipped", 
                 "description": f"Missing columns: {missing}", 
                 "details": {}
             }

        # Normalize Columns
        combined_norm = self.df['Combined'].astype(str).str.upper()
        type_of_program_norm = self.df[TYPE_COL_NAME].astype(str).str.lower()
        market_norm = self.df['Market'].astype(str).str.upper().str.strip()
        
        # Normalize Team Columns (Check for non-empty string content)
        home_team_valid = self.df['Home Team'].astype(str).str.strip().str.upper()
        away_team_valid = self.df['Away Team'].astype(str).str.strip().str.upper()
        
        # --- FILTERS ---

        # 1. Market Filter
        market_mask = market_norm == TARGET_MARKET.upper()
        
        # 2. Keyword Filter
        keyword_mask = combined_norm.str.contains(re.escape(KEYWORD), na=False)
        
        # 3. Team Validity Filter (NEW REQUIREMENT)
        # Must have valid content (not NaN, not empty, not 'NAN')
        teams_present_mask = (home_team_valid != 'NAN') & (home_team_valid != '') & \
                             (away_team_valid != 'NAN') & (away_team_valid != '')

        # Combine Filters: Match L/T keyword AND Market AND Teams are present
        target_rows_mask = keyword_mask & market_mask & teams_present_mask
        
        # --- Apply Logic ---
        
        # Apply the recommended status to valid target rows
        self.df.loc[target_rows_mask, FLAG_COLUMN] = 'Recommended: Live'

        # Check for anomalies (Target met but status is NOT Live)
        is_not_live_mask = type_of_program_norm != 'live'
        anomalous_flag_mask = target_rows_mask & is_not_live_mask
        
        self.df.loc[anomalous_flag_mask, FLAG_COLUMN] = 'ANOMALY: Should be Live (L/T Present & Teams Defined)'
        
        return {
            "check_key": "impute_lt_live_status",
            "status": "Flagged" if anomalous_flag_mask.sum() > 0 else "Completed",
            "description": f"Audited {KEYWORD} status in {TARGET_MARKET} (with Team Check). Flagged {anomalous_flag_mask.sum()} anomalies.",
            "details": {"rows_flagged": int(anomalous_flag_mask.sum()), "market": TARGET_MARKET, "keyword": KEYWORD}
        }

    def _consolidate_gillette_soccer_programs(self) -> Dict[str, Any]:
        """
        Identifies sequential 'Gillete Soccer' programs where the gap between the 
        End Time of the first and the Start Time of the second is 30 minutes or less.
        
        RESTRICTION: Applies ONLY to 'United Kingdom' and 'Ireland' markets.
        The second, later row is flagged for consolidation with a reference to the preceding row.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Consolidate_Gillete_Soccer'
        
        # --- ACCESS CONFIGURATION (or use defaults) ---
        # Using internal defaults here based on the prompt, but can use self.config if preferred
        KEYWORD = "GILLETE SOCCER"
        MAX_GAP_MINUTES = 30
        TARGET_MARKETS = ['UNITED KINGDOM', 'UK', 'IRELAND'] # <--- NEW RESTRICTION

        REQUIRED_COLS = ['Combined', 'Date', 'Start', 'End', 'Market', 'TV-Channel']
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "consolidate_gillete_soccer", "status": "Skipped",
                "action": "Program Consolidation Check", 
                "description": "Skipped: Missing required BSR columns.",
                "details": {"rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'
        
        # 1. Prepare Data & Timestamps
        try:
            # Normalize columns for filtering
            combined_norm = self.df['Combined'].astype(str).str.upper()
            market_norm = self.df['Market'].astype(str).str.strip().str.upper()
            
            # Create robust datetime objects
            date_key = self.df['Date'].astype(str).str[:10]
            self.df['Start_DT'] = pd.to_datetime(date_key + ' ' + self.df['Start'].astype(str), errors='coerce')
            
            # Adjust End_DT for midnight crossover
            end_times = self.df['End'].astype(str)
            base_end_dt = pd.to_datetime(date_key + ' ' + end_times, errors='coerce')
            rollover_mask = (base_end_dt < self.df['Start_DT']) & base_end_dt.notna()
            base_end_dt.loc[rollover_mask] += timedelta(days=1)
            self.df['End_DT'] = base_end_dt
            
        except Exception as e:
            return {
                "check_key": "consolidate_gillete_soccer", "status": "Failed",
                "action": "Program Consolidation Check", 
                "description": f"Failed to parse Date/Time columns: {e}",
                "details": {"rows_flagged": 0}
            }

        # 2. Create Filters
        # Filter A: Keyword Match
        gillete_mask = combined_norm.str.contains(re.escape(KEYWORD), na=False)
        
        # Filter B: Market Match (UK or Ireland) <--- NEW
        market_mask = market_norm.isin(TARGET_MARKETS)
        
        # Filter C: Valid Times
        valid_time_mask = self.df['Start_DT'].notna() & self.df['End_DT'].notna()

        # 3. Filter and Sort Candidates
        # Combine all masks to select rows
        df_candidates = self.df[gillete_mask & market_mask & valid_time_mask].copy()
        
        df_candidates['Original_Index'] = df_candidates.index
        GROUP_COLS = ['Market', 'TV-Channel']
        
        # Sort globally by grouping cols and time to ensure correct sequence
        df_candidates = df_candidates.sort_values(by=GROUP_COLS + ['Start_DT'])

        # 4. Perform Sequential Gap Check
        complex_flags = {}
        
        for _, group in df_candidates.groupby(GROUP_COLS):
            # Calculate gap: Start of Current - End of Previous
            time_gap = (group['Start_DT'] - group['End_DT'].shift(1)) / timedelta(minutes=1)
            
            preceding_original_indices = group['Original_Index'].shift(1)
            
            # Identify consolidation candidates (gap is positive but small)
            consolidation_mask = (time_gap <= MAX_GAP_MINUTES) & (time_gap >= 0)
            
            indices_now = group[consolidation_mask]['Original_Index']
            preceding_indices = group['Original_Index'].shift(1)[consolidation_mask]

            # Construct detailed flags
            for curr_idx, prev_idx in zip(indices_now, preceding_indices):
                # Use .iloc[0] on the looked-up value to ensure we get a scalar string
                prev_start_val = self.df.loc[prev_idx, 'Start']
                
                msg = (f"Consolidate with program starting at {prev_start_val} "
                       f"(Original Index: {int(prev_idx)}, Gap <= {MAX_GAP_MINUTES}min)")
                complex_flags[curr_idx] = msg

        rows_flagged = len(complex_flags)
        
        # 5. Apply Flag to Original DataFrame
        if rows_flagged > 0:
            flag_series = pd.Series(complex_flags)
            self.df.loc[flag_series.index, FLAG_COLUMN] = flag_series

        # Final cleanup
        self.df.drop(columns=['Start_DT', 'End_DT'], inplace=True, errors='ignore')

        return {
            "check_key": "consolidate_gillete_soccer",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Program Consolidation Check", 
            "description": f"Flagged {rows_flagged} sequential '{KEYWORD}' rows for consolidation in UK/Ireland (gap <= {MAX_GAP_MINUTES} min).",
            "details": {
                "rows_flagged": int(rows_flagged),
                "max_gap_minutes": MAX_GAP_MINUTES,
                "target_keyword": KEYWORD,
                "target_markets": TARGET_MARKETS
            }
        }

    def _check_sky_showcase_live_status(self) -> Dict[str, Any]:
        """
        Implements a zero-tolerance check: flags any program on 'Sky Showcase' 
        in the UK/United Kingdom market that is incorrectly labeled 'Live', 
        as this channel is designated for Repeat/Delayed content only.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Sky_Showcase_Live_Flag'
        
        # Define specific target parameters using robust strings/regex
        TARGET_MARKET_REGEX = r'UNITED KINGDOM|UK' # Broadening to include full name and Ireland (common data area)
        TARGET_CHANNEL_KEYWORD = 'SKY SHOWCASE' # Keyword in the TV-Channel name
        FORBIDDEN_STATUS = 'LIVE'
        
        REQUIRED_COLS = ['Market', 'TV-Channel', 'Type of program']
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "check_sky_showcase_live", "status": "Skipped",
                "action": "Zero-Tolerance Live Check", 
                "description": "Skipped: Missing required BSR columns.",
                "details": {"rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'

        # 1. Normalize columns for reliable filtering
        market_norm = self.df['Market'].astype(str).str.strip().str.upper()
        channel_norm = self.df['TV-Channel'].astype(str).str.strip().str.upper()
        type_norm = self.df['Type of program'].astype(str).str.strip().str.upper()

        # 2. Identify the target rows (Robust Market AND Channel Identification)
        
        # Mask 1: Identify UK/Ireland market variants
        market_match_mask = market_norm.str.contains(TARGET_MARKET_REGEX, regex=True, na=False)
        
        # Mask 2: Identify Sky Showcase variants (uses str.contains to catch "Sky Showcase DE")
        channel_match_mask = channel_norm.str.contains(TARGET_CHANNEL_KEYWORD, na=False)
        
        target_rows_mask = market_match_mask & channel_match_mask
        
        # 3. Identify the error condition (Target row AND status is 'LIVE')
        error_mask = target_rows_mask & (type_norm == FORBIDDEN_STATUS)
        
        rows_flagged = error_mask.sum()
        
        # 4. Apply Flag to Original DataFrame
        if rows_flagged > 0:
            
            # Construct the flag message
            flag_message = f"INTEGRITY ERROR: Designated repeat channel ({TARGET_CHANNEL_KEYWORD} variant) is incorrectly marked '{FORBIDDEN_STATUS}'."
            
            # Apply flag only to rows currently marked OK
            rows_to_flag = error_mask & (self.df[FLAG_COLUMN] == 'OK')
            
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message

        # 5. Final Summary
        return {
            "check_key": "check_sky_showcase_live",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Zero-Tolerance Live Check", 
            "description": f"Flagged {rows_flagged} rows on Sky Showcase variants that were incorrectly tagged as 'Live'.",
            "details": {
                "rows_flagged": int(rows_flagged),
                "target_channel_keyword": TARGET_CHANNEL_KEYWORD,
                "forbidden_status": FORBIDDEN_STATUS
            }
        }

    def _standardize_uk_ire_region(self) -> Dict[str, Any]:
        """
        Audits Region Standardization for UK and Ireland markets.
        
        Checks:
        1. Ireland Check: Flags 'Ireland' programs incorrectly labeled with 'United Kingdom' region (Should be 'Europe').
        2. UK Check: Flags 'United Kingdom' programs incorrectly labeled with 'Europe' region (Should be 'United Kingdom').
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Region_Standardization_Flag'
        
        REQUIRED_COLS = ['Market', 'Region']
        
        # Check for required columns
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "audit_uk_ire_region", "status": "Skipped",
                "action": "Region Standardization", 
                "description": "Skipped: Missing required columns.",
                "details": {"rows_flagged": 0}
            }

        # Initialize flag column if it doesn't exist, otherwise keep existing flags
        if FLAG_COLUMN not in self.df.columns:
            self.df[FLAG_COLUMN] = 'OK'

        # 1. Prepare normalized Market and Region columns
        # Remove punctuation to ensure 'IRELAND' matches 'Ireland.' or 'Ireland'
        market_norm = self.df['Market'].astype(str).str.upper()
        market_norm = market_norm.str.replace(r'[^A-Z\s]', '', regex=True).str.strip()
        
        region_norm = self.df['Region'].astype(str).str.upper()
        region_norm = region_norm.str.replace(r'[^A-Z\s]', '', regex=True).str.strip()
        
        # 2. Define Groups
        
        # Group A: Ireland Markets
        is_ireland_market = market_norm == 'IRELAND'
        
        # Group B: UK Markets (United Kingdom, UK, Great Britain)
        uk_variations = ['UNITED KINGDOM', 'UK', 'GREAT BRITAIN']
        is_uk_market = market_norm.isin(uk_variations)
        
        # 3. Identify Inconsistencies
        
        # Error 1: Market is IRELAND, but Region is UK (Should be Europe)
        # We detect if region is one of the UK variations
        is_uk_region = region_norm.isin(uk_variations)
        error_ireland = is_ireland_market & is_uk_region
        
        # Error 2: Market is UK, but Region is EUROPE (Should be UK)
        is_europe_region = region_norm == 'EUROPE'
        error_uk = is_uk_market & is_europe_region
        
        # 4. Apply Flags
        
        # Flag Ireland Errors
        if error_ireland.any():
            flag_message = "INCORRECT REGION: Ireland program detected with UK Region. Should be 'Europe'."
            rows_to_flag = error_ireland & (self.df[FLAG_COLUMN] == 'OK')
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message
            
        # Flag UK Errors
        if error_uk.any():
            flag_message = "INCORRECT REGION: UK program detected with 'Europe' Region. Should be 'United Kingdom'."
            rows_to_flag = error_uk & (self.df[FLAG_COLUMN] == 'OK')
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message

        # Calculate Total Flagged
        # Note: Using bitwise OR to count unique rows flagged by either condition
        total_rows_flagged = (error_ireland | error_uk).sum()

        # 5. Final Summary
        return {
            "check_key": "audit_uk_ire_region",
            "status": "Flagged" if total_rows_flagged > 0 else "Completed",
            "action": "Region Standardization", 
            "description": f"Flagged {total_rows_flagged} rows with Region/Market inconsistencies.",
            "details": {
                "rows_processed": int(initial_rows),
                "total_flagged": int(total_rows_flagged),
                "ireland_errors": int(error_ireland.sum()),
                "uk_errors": int(error_uk.sum())
            }
        }

    def _check_fixture_vs_case(self) -> Dict[str, Any]:
        """
        Checks the 'Phase / Fixture / Episode Desc.' column for incorrect casing of 
        the separator 'VS' (must be lowercase 'vs') and flags violating rows.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Fixture_Vs_Case_Flag'
        TARGET_COL = 'Phase / Fixture / Episode Desc.'
        
        # Target the specific market for the check (UK Market only)
        TARGET_MARKET = 'United Kingdom'
        
        REQUIRED_COLS = ['Market', TARGET_COL]
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "check_fixture_vs_case", "status": "Skipped",
                "action": "Fixture Case Check", 
                "description": "Skipped: Missing required BSR columns.",
                "details": {"rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'
        
        # 1. Normalize market column for filtering
        market_norm = self.df['Market'].astype(str).str.strip().str.upper()
        uk_market_mask = market_norm == TARGET_MARKET.upper()
        
        # 2. Identify the error condition: Uppercase 'VS' variants
        
        # We must look for any instance of 'VS' that is NOT entirely lowercase 'vs'
        # The safest way is to find non-lowercase instances of V/S surrounded by spaces, or check if the title contains V or S.
        
        # Mask A: Rows in the target market
        
        # Mask B: Rows that contain the uppercase form of 'VS' (case-sensitive check for the error)
        # We look for "VS", "Vs", or "V.S." (space-sensitive) in the column content.
        # Note: We must ensure we don't flag words that start with V or S.
        
        # Check 1: Find rows that contain 'VS' or 'Vs' (must be applied to the column content)
        target_content = self.df[TARGET_COL].astype(str).str.strip()
        
        # Use regex to find "VS" or "Vs" surrounded by non-word boundaries or spaces
        # We will use the simple regex pattern (VS or Vs) within word boundaries or spaces
        
        # A robust way to check for the improper casing:
        # 1. Standardize the whole column to lowercase.
        # 2. Check the difference between the original and the standardized column where the substring 'vs' is present.
        
        # Filter only UK market rows for analysis
        target_content_uk = target_content[uk_market_mask].copy()
        
        # Check 1: Does the content contain any form of 'VS' (case-insensitive)?
        vs_present_mask = target_content_uk.str.contains(r'VS', case=False, na=False)
        
        # Check 2: Does the content contain the invalid uppercase form?
        # We must specifically check for V or S being capitalized in the context of 'vs'.
        
        # To check for improper capitalization, we look for 'VS' or 'Vs' in the original content
        improper_case_mask = target_content_uk.str.contains(r'(VS|Vs|V\s+S)', case=True, na=False)
        
        # Final error mask: In the UK market, VS is present, AND the casing is improper (non-lowercase)
        error_mask = uk_market_mask & improper_case_mask.reindex(self.df.index).fillna(False)
        
        rows_flagged = error_mask.sum()
        
        # 3. Apply Flag
        if rows_flagged > 0:
            
            flag_message = "CASE INTEGRITY ERROR: Fixture must use only lowercase 'vs'. Uppercase ('VS' or 'Vs') found."
            
            # Apply flag only to rows currently marked OK
            rows_to_flag = error_mask & (self.df[FLAG_COLUMN] == 'OK')
            
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message

        # 4. Final Summary
        return {
            "check_key": "check_fixture_vs_case",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Fixture Case Check", 
            "description": f"Flagged {rows_flagged} UK rows where the fixture delimiter was incorrectly capitalized (must be 'vs').",
            "details": {
                "rows_flagged": int(rows_flagged),
                "target_market": TARGET_MARKET
            }
        }

    def _check_pan_balkans_serbia_parity(self) -> Dict[str, Any]:
        """
        Checks if the total program row count for 'Pan Balkans' is strictly equal to 
        the total program row count for 'Serbia', enforcing structural parity post-modeling.
        Correctly handles hyphens/spaces in market names.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Balkans_Serbia_Parity_Flag'
        
        # Define Target Markets (Use the clean form without spaces/hyphens)
        TARGET_MARKET_A_CLEAN = 'PANBALKANS' # Pan-Balkans
        TARGET_MARKET_B_CLEAN = 'SERBIA' # Serbia
        
        REQUIRED_COLS = ['Market']
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "check_pan_balkans_serbia_parity", "status": "Skipped",
                "action": "Market Parity Check", 
                "description": "Skipped: Missing required BSR 'Market' column.",
                "details": {"parity_match": "False", "rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'
        
        # 1. Normalize market column: Upper, strip, and remove all hyphens and spaces
        market_norm = self.df['Market'].astype(str).str.upper().str.strip()
        market_norm = market_norm.str.replace(r'[\s\-]+', '', regex=True) # <-- FIX: Removes spaces and hyphens

        # 2. Count Total Programs in each target market
        
        # Count rows in Market A (Matching against the clean constant)
        count_a = (market_norm == TARGET_MARKET_A_CLEAN).sum()
        
        # Count rows in Market B
        count_b = (market_norm == TARGET_MARKET_B_CLEAN).sum()
        
        # 3. Perform Parity Check
        is_parity_match = (count_a > 0) and (count_a == count_b)
        
        rows_flagged = 0
        
        if not is_parity_match:
            
            # Apply the mismatch mask based on the normalized values
            mismatch_mask = (market_norm == TARGET_MARKET_A_CLEAN) | (market_norm == TARGET_MARKET_B_CLEAN)
            
            # Apply the flag (using the assumed clean original names for the message)
            flag_message = (f"PARITY ERROR: Program count mismatch between Pan Balkans ({count_a} rows) "
                            f"and Serbia ({count_b} rows). Counts must be identical.")
            
            rows_to_flag = mismatch_mask & (self.df[FLAG_COLUMN] == 'OK')
            
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message
            rows_flagged = rows_to_flag.sum()


        # 4. Final Summary
        return {
            "check_key": "check_pan_balkans_serbia_parity",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Market Parity Check", 
            "description": f"Audited program count parity. Counts: Pan Balkans ({count_a}), Serbia ({count_b}).",
            "details": {
                "rows_flagged": int(rows_flagged),
                "pan_balkans_count": int(count_a),
                "serbia_count": int(count_b),
                "parity_match": str(is_parity_match)
            }
        }

    def _audit_multi_match_status(self) -> Dict[str, Any]:
        """
        Audits Multi-Match content and Classification Codes.
        
        Logic:
        1. Uses 'Combined' column exclusively for content checks.
        2. Check A: If 'Combined' contains 'Goal Rush'/'Konferenz', 'Phase / Fixture / Episode Desc.' must have 'MULTIMATCH' tag.
        3. Check B: If 'Combined' ends in 'NB' or 'VB', 'Type of program' must be 'Magazine & Support'.
        4. Check C: If tagged 'MULTIMATCH', 'Type of program' must be 'Live' or 'Repeat'.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Multi_Match_Audit_Flag'
        
        # Define keywords
        MULTI_MATCH_KEYWORDS = ['GOAL RUSH', 'KONFERENZ', 'CONFERENCE']
        VALID_MULTIMATCH_TYPES = ['LIVE', 'REPEAT']
        
        # Define Columns
        COMBINED_COL = 'Combined'
        FIXTURE_DESC_COL = 'Phase / Fixture / Episode Desc.'
        TYPE_COL = 'Type of programme'
        
        # Regex for valid tag and suffixes
        EXPECTED_FIXTURE_REGEX = r'MULTI[\s\-]*MATCH' 
        # Suffix regex: matches space followed by NB or VB at the end of the string
        SUFFIX_REGEX = r'\s(NB|VB)$'  
        
        # Strict check for required columns
        REQUIRED_COLS = [COMBINED_COL, FIXTURE_DESC_COL, TYPE_COL]
        
        if not all(col in self.df.columns for col in REQUIRED_COLS):
             return {
                "check_key": "audit_multi_match", "status": "Skipped",
                "action": "Multi-Match & Code Audit", 
                "description": f"Skipped: Missing required columns (Need '{COMBINED_COL}', '{FIXTURE_DESC_COL}', '{TYPE_COL}').",
                "details": {"rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'

        # 1. Prepare normalized columns
        combined_norm = self.df[COMBINED_COL].astype(str).str.upper().fillna('')
        fixture_desc_norm = self.df[FIXTURE_DESC_COL].astype(str).str.upper().fillna('')
        type_norm = self.df[TYPE_COL].astype(str).str.strip().str.upper()
        
        # 2. Define Condition Masks
        
        # A: Combined contains Multi-Match Keyword (e.g., "Goal Rush")
        match_keyword_pattern = '|'.join([re.escape(k) for k in MULTI_MATCH_KEYWORDS])
        combined_has_keyword = combined_norm.str.contains(match_keyword_pattern, na=False)
        
        # B: Fixture Column contains "MULTIMATCH" Tag
        tag_is_present = fixture_desc_norm.str.contains(EXPECTED_FIXTURE_REGEX, regex=True, na=False)
        
        # C: Combined ends with NB or VB
        # This regex looks for whitespace + NB/VB + end of string
        ends_with_code = combined_norm.str.contains(SUFFIX_REGEX, regex=True, na=False)
        
        # D: Type Checks
        type_is_live_repeat = type_norm.isin(VALID_MULTIMATCH_TYPES)
        type_is_magazine = type_norm == 'MAGAZINE & SUPPORT'

        # 3. Identify Errors
        
        # Error 1: MISSING MULTI-MATCH TAG
        # Logic: If Combined has Keyword (Goal Rush) -> Must have MULTIMATCH tag.
        # EXCEPTION: If it is a Magazine (NB/VB), it might be "Goal Rush NB" which is a support show.
        missing_tag_mask = (combined_has_keyword & ~ends_with_code) & (~tag_is_present)
        
        # Error 2: INVALID TYPE FOR NB/VB SUFFIX (The critical check requested)
        # Logic: If Combined ends in NB or VB -> Type MUST be 'Magazine & Support'
        invalid_magazine_mask = ends_with_code & (~type_is_magazine)
        
        # Error 3: INVALID TYPE FOR MULTIMATCH TAG
        # Logic: If explicitly tagged MULTIMATCH -> Type MUST be Live or Repeat
        invalid_multimatch_type_mask = tag_is_present & (~type_is_live_repeat)

        # 4. Apply Flags (Priority Order)
        
        # Apply Error 2 (NB/VB Type Mismatch) - High Priority
        if invalid_magazine_mask.any():
            msg = f"TYPE MISMATCH: Combined description ends in NB/VB, so '{TYPE_COL}' must be 'Magazine & Support'."
            rows_to_flag = invalid_magazine_mask & (self.df[FLAG_COLUMN] == 'OK')
            self.df.loc[rows_to_flag, FLAG_COLUMN] = msg

        # Apply Error 1 (Missing Multi-Match Tag)
        if missing_tag_mask.any():
            msg = f"FIXTURE TAG MISSING: Content indicates Multi-Match, but '{FIXTURE_DESC_COL}' is missing 'MULTIMATCH' tag."
            rows_to_flag = missing_tag_mask & (self.df[FLAG_COLUMN] == 'OK')
            self.df.loc[rows_to_flag, FLAG_COLUMN] = msg

        # Apply Error 3 (Invalid Multi-Match Type)
        if invalid_multimatch_type_mask.any():
            msg = f"INVALID TYPE: Item is tagged 'MULTIMATCH', but '{TYPE_COL}' is not Live or Repeat."
            rows_to_flag = invalid_multimatch_type_mask & (self.df[FLAG_COLUMN] == 'OK')
            self.df.loc[rows_to_flag, FLAG_COLUMN] = msg

        # Recalculate total flagged
        total_flagged = (self.df[FLAG_COLUMN] != 'OK').sum()

        # 5. Final Summary
        return {
            "check_key": "audit_multi_match",
            "status": "Flagged" if total_flagged > 0 else "Completed",
            "action": "Multi-Match & Code Audit", 
            "description": f"Audited Multi-Match tags and NB/VB codes using 'Combined' column. Flagged {total_flagged} rows.",
            "details": {
                "rows_processed": int(initial_rows),
                "nb_vb_errors": int(invalid_magazine_mask.sum()),
                "missing_tag_errors": int(missing_tag_mask.sum()),
                "invalid_type_errors": int(invalid_multimatch_type_mask.sum())
            }
        }
    
    def _check_date_time_format_integrity(self) -> Dict[str, Any]:
        """
        Audits specific date and time columns to check for data type inconsistencies 
        (e.g., numeric entries, invalid formats) that prevent UTC format conversion.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_DateTime_Format_Flag'
        
        # Define all six columns that need to be checked
        DATE_TIME_COLS_TO_CHECK = [
            'Date (UTC/GMT)', 'Date', 
            'Start (UTC)', 'End (UTC)', 
            'Start', 'End'
        ]
        
        self.df[FLAG_COLUMN] = 'OK'
        
        # 1. Check for required column existence
        if not all(col in self.df.columns for col in DATE_TIME_COLS_TO_CHECK):
            missing = list(set(DATE_TIME_COLS_TO_CHECK) - set(self.df.columns))
            return {
                "check_key": "check_datetime_format", "status": "Skipped",
                "action": "Date/Time Integrity Check", 
                "description": f"Skipped: Missing required columns: {missing}",
                "details": {"rows_flagged": 0}
            }

        # --- Data Type Check Loop ---
        
        total_flagged_rows = 0
        
        for col in DATE_TIME_COLS_TO_CHECK:
            
            # Determine if the column contains Date/DateTime or Time/Duration data
            if 'Start' in col or 'End' in col or 'Duration' in col:
                # For Time/Duration columns, check if they can be converted to timedelta
                # We must convert the column to string first to handle numeric entries like '11232'.
                try:
                    # pandas' to_timedelta can handle HH:MM:SS or large numbers (Excel format)
                    parsed_series = pd.to_timedelta(self.df[col].astype(str), errors='coerce')
                    error_mask = parsed_series.isna()
                except Exception:
                    # Fallback check if to_timedelta raises unexpected error
                    error_mask = self.df[col].astype(str).str.contains(r'[A-Za-z]', na=False) # Check for letters
            else:
                # For Date columns, check if they can be converted to datetime
                try:
                    # pandas' to_datetime can handle various date formats
                    parsed_series = pd.to_datetime(self.df[col], errors='coerce')
                    error_mask = parsed_series.isna()
                except Exception:
                    # Fallback check for general corruption
                    error_mask = self.df[col].astype(str).str.contains(r'[A-Za-z]', na=False) 
            
            
            # 2. Apply Flag to the BSR
            if error_mask.any():
                
                flag_message = f"FORMAT ERROR: '{col}' contains invalid or non-standard entries (e.g., numeric IDs, text). Requires manual cleanup."
                
                # Identify rows that failed the current check AND were not already flagged
                rows_to_flag = error_mask & (self.df[FLAG_COLUMN] == 'OK')
                
                self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message
                total_flagged_rows += rows_to_flag.sum()


        # 3. Final Summary
        return {
            "check_key": "check_datetime_format",
            "status": "Flagged" if total_flagged_rows > 0 else "Completed",
            "action": "Date/Time Integrity Check", 
            "description": f"Audited {len(DATE_TIME_COLS_TO_CHECK)} date/time columns. Flagged {total_flagged_rows} entries with invalid formatting.",
            "details": {
                "rows_flagged": int(total_flagged_rows),
                "columns_checked": DATE_TIME_COLS_TO_CHECK
            }
        }

    def _check_live_broadcast_uniqueness(self) -> Dict[str, Any]:
            """
            Ensures that a specific live broadcast slot (defined by Market, TV-Channel, 
            and the specific Fixture Description) has no time overlap with itself 
            (i.e., checks for duplicate Live entries for the same match).
            """
            initial_rows = len(self.df)
            FLAG_COLUMN = 'QC_Live_Overlap_Flag'
            
            # --- Define the Grouping Key ---
            # We group by Market, Channel, and the specific Fixture description.
            GROUPING_KEY_COLS = ['Market', 'TV-Channel', 'Phase / Fixture / Episode Desc.']
            
            LIVE_PROGRAM_TYPE = 'LIVE'
            
            # FIX: Updated column name to match your file ('Type of programme')
            TYPE_COL = 'Type of programme' 
            
            REQUIRED_COLS = GROUPING_KEY_COLS + [TYPE_COL, 'Date (UTC/GMT)', 'Start (UTC)', 'End (UTC)']
            
            # DEBUG: Print missing columns if any
            missing = [c for c in REQUIRED_COLS if c not in self.df.columns]
            if missing:
                return {
                    "check_key": "check_live_broadcast_uniqueness", "status": "Skipped", 
                    "action": "Live Overlap Check", 
                    "description": f"Skipped: Missing columns: {missing}", 
                    "details": {"rows_flagged": 0}
                }

            self.df[FLAG_COLUMN] = 'OK'
            
            # --- 1. Prepare Data and Timestamps ---
            try:
                date_key = self.df['Date (UTC/GMT)'].astype(str).str[:10]
                self.df['Start_DT'] = pd.to_datetime(date_key + ' ' + self.df['Start (UTC)'].astype(str), errors='coerce')
                base_end_dt = pd.to_datetime(date_key + ' ' + self.df['End (UTC)'].astype(str), errors='coerce')
                
                # Handle midnight rollover
                rollover_mask = (base_end_dt < self.df['Start_DT']) & base_end_dt.notna()
                base_end_dt.loc[rollover_mask] += timedelta(days=1)
                self.df['End_DT'] = base_end_dt
                
            except Exception as e:
                self.df.drop(columns=['Start_DT', 'End_DT'], inplace=True, errors='ignore')
                return {"check_key": "check_live_broadcast_uniqueness", "status": "Failed", "action": "Live Overlap Check", "description": f"Failed to parse Date/Time columns: {e}", "details": {"rows_flagged": 0}}
                
            # Standardize grouping columns
            for col in GROUPING_KEY_COLS:
                self.df[col] = self.df[col].astype(str).str.strip().str.upper().str.replace(r'[^A-Z0-9\s\.\-]', '', regex=True).fillna('NAN')

            # Filter for LIVE programs only (Using the correct column name)
            live_mask = self.df[TYPE_COL].astype(str).str.upper().str.strip() == LIVE_PROGRAM_TYPE
            
            df_live_candidates = self.df[live_mask].copy()
            
            # Sort by Grouping Key + Start Time
            df_live_candidates = df_live_candidates.sort_values(by=GROUPING_KEY_COLS + ['Start_DT'])
            
            # --- 2. Overlap Detection Logic ---
            
            conflict_details = {} 
            
            # Group by the 3-part key
            for key_tuple, group in df_live_candidates.groupby(GROUPING_KEY_COLS):
                if len(group) < 2:
                    continue
                    
                lagged_end_dt = group['End_DT'].shift(1)
                
                # Overlap occurs if Start Time < Previous End Time
                overlap_start_mask = group['Start_DT'] < lagged_end_dt
                
                if overlap_start_mask.any():
                    
                    current_overlap_indices = group[overlap_start_mask].index.tolist()
                    preceding_overlap_indices = group[overlap_start_mask].shift(1).index.dropna().astype(int)

                    all_conflict_indices = set(current_overlap_indices).union(set(preceding_overlap_indices))
                    
                    # Format the detailed conflict message
                    conflict_log = []
                    
                    for idx in sorted(list(all_conflict_indices)):
                        row = self.df.loc[idx]
                        log_entry = (f"Index {idx} | "
                                    f"Date: {row['Date (UTC/GMT)'].strftime('%Y-%m-%d')} | "
                                    f"Times: {row['Start (UTC)']} - {row['End (UTC)']}")
                        conflict_log.append(log_entry)
                    
                    key_id = "|".join([str(k) for k in key_tuple]) 
                    conflict_message = f"DUPLICATION ERROR: Fixture '{key_id}' has overlapping LIVE entries. Rows: " + " || ".join(conflict_log)

                    for idx in all_conflict_indices:
                        conflict_details[idx] = conflict_message
            

            # --- 3. Apply Flag to Original DataFrame ---
            rows_flagged = len(conflict_details)
            
            if rows_flagged > 0:
                flag_series = pd.Series(conflict_details)
                self.df.loc[flag_series.index, FLAG_COLUMN] = flag_series

            # Final cleanup
            self.df.drop(columns=['Start_DT', 'End_DT'], inplace=True, errors='ignore')

            # 4. Final Summary
            return {
                "check_key": "check_live_broadcast_uniqueness",
                "status": "Flagged" if rows_flagged > 0 else "Completed",
                "action": "Live Uniqueness Check", 
                "description": f"Flagged {rows_flagged} rows where the same fixture was listed as Live multiple times overlapping.",
                "details": {
                    "rows_flagged": int(rows_flagged),
                    "grouping_keys": GROUPING_KEY_COLS
                }
            }
    
    def _audit_channel_line_item_count(self) -> Dict[str, Any]:
        """
        Calculates the total number of line items (programs) for each unique TV-Channel 
        in the BSR and returns this summary as a separate DataFrame for reporting.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Channel_Count_Audit_Flag'
        
        CHANNEL_COL = 'TV-Channel'
        
        if CHANNEL_COL not in self.df.columns:
            return {
                "check_key": "audit_channel_line_item_count", "status": "Skipped",
                "action": "Deliverable Count Audit", 
                "description": "Skipped: Missing required BSR 'TV-Channel' column.",
                "details": {"report_generated": False}
            }

        # Initialize the flag column for audit (optional, but good practice)
        self.df[FLAG_COLUMN] = 'OK' 

        # 1. Normalize and Calculate Counts
        
        # Normalize channel names for accurate grouping (UPPER/strip)
        channel_norm = self.df[CHANNEL_COL].astype(str).str.strip().str.upper()
        
        # Calculate the current line item count for each unique channel
        channel_counts_df = channel_norm.value_counts().reset_index()
        channel_counts_df.columns = ['TV-Channel_Norm', 'Program_Count']
        
        # Sort for better readability in the final report
        channel_counts_df = channel_counts_df.sort_values(by='Program_Count', ascending=False)
        
        # 2. Final Summary
        return {
            "check_key": "audit_channel_line_item_count",
            "status": "Completed",
            "action": "Deliverable Count Audit", 
            "description": f"Generated line item count summary for {len(channel_counts_df)} unique channels.",
            "details": {
                "total_channels": int(len(channel_counts_df)),
                "report_generated": True,
                # CRITICAL: Return the DataFrame itself for saving to a separate tab
                "channel_count_report_df": channel_counts_df.to_dict('records')
            }
        }

    def _check_combined_archive_status(self) -> Dict[str, Any]:
        """
        Audits the 'Combined' column for the keyword 'archive' and flags rows 
        as potential archival content requiring removal or review.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Archive_Status_Flag'
        KEYWORD = 'ARCHIVE'
        COMBINED_COL = 'Combined'
        
        # Check for required column
        if COMBINED_COL not in self.df.columns:
            return {
                "check_key": "check_combined_archive_status", "status": "Skipped",
                "action": "Keyword Audit", 
                "description": "Skipped: Missing required BSR 'Combined' column.",
                "details": {"rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'

        # 1. Normalize the Combined column for case-insensitive search
        combined_norm = self.df[COMBINED_COL].astype(str).str.upper()
        
        # 2. Identify rows containing the keyword
        archive_mask = combined_norm.str.contains(KEYWORD, na=False)
        
        rows_flagged = archive_mask.sum()
        
        # 3. Apply Flag to Original DataFrame
        if rows_flagged > 0:
            
            flag_message = f"ARCHIVAL CONTENT FLAG: Keyword '{KEYWORD}' found in Combined column. Requires review/removal."
            
            # Apply flag only to rows currently marked OK
            rows_to_flag = archive_mask & (self.df[FLAG_COLUMN] == 'OK')
            
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message

        # 4. Final Summary
        return {
            "check_key": "check_combined_archive_status",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Keyword Audit", 
            "description": f"Flagged {rows_flagged} rows containing the '{KEYWORD}' keyword in the Combined column.",
            "details": {
                "rows_flagged": int(rows_flagged),
                "target_keyword": KEYWORD
            }
        }

    def _suppress_duplicated_audience(self) -> Dict[str, Any]:
        """
        Audits the BSR to flag any row where the 'Source' column indicates a duplication 
        origin, but EITHER the Modeled Audience or the Metered Audience column contains 
        a positive, non-zero value (an anomaly).
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Audience_Suppression_Flag'
        SOURCE_COL = 'Source'
        KEYWORD = 'DUPLICATED FROM BSA'
        
        # Define BOTH audience columns that must be zero
        TARGET_AUDIENCE_COLS = [
            'Aud. Estimates [\'000s]', 
            'Aud Metered (000s) 3+'
        ]
        
        # Check for required columns
        REQUIRED_COLS = TARGET_AUDIENCE_COLS + [SOURCE_COL]
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "suppress_duplicated_audience", "status": "Skipped",
                "action": "Audience Suppression Audit", 
                "description": "Skipped: Missing required BSR columns.",
                "details": {"rows_flagged": 0}
            }

        self.df[FLAG_COLUMN] = 'OK'

        # 1. Identify rows that are duplication sources
        source_norm = self.df[SOURCE_COL].astype(str).str.upper()
        suppression_mask = source_norm.str.contains(KEYWORD, na=False)
        
        # 2. Identify the anomaly: Is there ANY positive audience value?
        
        # Check if ANY of the two target columns are greater than zero
        # Use fillna(0) to treat NaNs as zero, ensuring a safe comparison
        audience_check_df = self.df[TARGET_AUDIENCE_COLS].fillna(0)
        
        # Create a mask that is TRUE if AT LEAST ONE of the two columns is positive
        any_audience_positive_mask = (audience_check_df > 0).any(axis=1)
        
        # Final Error Mask: Duplication Source AND Any Positive Audience
        error_mask = suppression_mask & any_audience_positive_mask
        
        rows_flagged = error_mask.sum()
        
        # 3. Apply Flag to Original DataFrame (No Value Update)
        if rows_flagged > 0:
            
            flag_message = f"SUPPRESSION ANOMALY: Source indicates duplication origin ('{KEYWORD}'), but Audience is POSITIVE in Aud. Estimates or Aud Metered."
            
            # Apply flag only to rows currently marked OK
            rows_to_flag = error_mask & (self.df[FLAG_COLUMN] == 'OK')
            
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message

        # 4. Final Summary
        return {
            "check_key": "suppress_duplicated_audience",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Audience Suppression Audit", 
            "description": f"Flagged {rows_flagged} duplication source rows containing positive audience values (should be zero).",
            "details": {
                "rows_flagged": int(rows_flagged),
                "target_columns_checked": TARGET_AUDIENCE_COLS
            }
        }

    # --- HELPER FUNCTION: CHANNEL CLASSIFICATION (Outside the main function) ---

    # def _get_uk_ire_channel_map(self):
    #     """
    #     Defines and processes the specific rules for mapping UK and Ireland channels 
    #     to their unique identifiers.
    #     """
    #     # Channel list and counts provided by the user
    #     RAW_CHANNELS_LIST = {
    #         'BBC1': 2, 'BBC2': 2, 'Channel 4': 2, 'Premier Sports 1 IRE': 1,
    #         'Premier Sports 1 UK': 1, 'Quest': 2, 'Sky Mix': 2, 'Sky News': 1,
    #         'Sky News UK': 1, 'Sky Showcase IRE': 1, 'Sky Showcase UK': 1, 
    #         'Sky Sports Football': 2, 'Sky Sports Golf': 1, 'Sky Sports Golf UK': 1,
    #         'Sky Sports Main Event': 2, 'Sky Sports News': 1, 'Sky Sports News UK': 1,
    #         'Sky Sports Premier League': 1, 'Sky Sports Premier League UK': 1,
    #         'Sky Sports Racing': 2, 'Sky Sports Tennis IRE': 1, 
    #         'Sky Sports Tennis UK': 1, 'TNT Sports 1': 2, 'U+Dave': 2
    #     }

    #     # Channels that default to Ireland if no suffix is found
    #     IRELAND_DEFAULT_CHANNELS = {
    #         'SKY NEWS', 'SKY SPORTS GOLF', 'SKY SPORTS NEWS', 'SKY SPORTS PREMIER LEAGUE',
    #     }

    #     # 1. Create a definitive mapping for (Normalized Channel Name, Market) -> Canonical Channel ID
    #     channel_map = {}
        
    #     for raw_name, count in RAW_CHANNELS_LIST.items():
    #         # Standardize the base name (remove suffixes for core identity)
    #         base_name = raw_name.replace(' IRE', '').replace(' UK', '').strip().upper()
            
    #         if count == 2:
    #             # Rule 1: Dual-Market Channels (Same name in both)
    #             # The core identity is the channel name itself (e.g., 'BBC1')
    #             channel_map[base_name] = base_name 
            
    #         elif count == 1:
    #             # Rule 2: Single-Market Channels (Requires suffix)
                
    #             # Extract market suffix if present
    #             market_suffix = 'IRE' if 'IRE' in raw_name else ('UK' if 'UK' in raw_name else None)
                
    #             # Check for the channels that default to Ireland if no suffix is found
    #             if not market_suffix and base_name in IRELAND_DEFAULT_CHANNELS:
    #                 market_suffix = 'IRE'
                
    #             if market_suffix:
    #                 canonical_id = f"{base_name}_{market_suffix}"
                    
    #                 # Check for the core channel name without the suffix in the map
    #                 if base_name not in channel_map:
    #                     channel_map[base_name] = canonical_id
                    
    #     return channel_map

    # def _safe_get_time_string(self, series):
    #     """Safely extracts time string from mixed data types, enforcing string conversion."""
        
    #     # CRITICAL FIX: Convert the series to string format FIRST to handle float/NaN corruption
    #     series_str = series.astype(str) 
        
    #     # Attempt to convert to datetime, coercing errors to NaT
    #     # Note: We must also handle non-time string data (like '00:00:00') that can result from NaT conversion.
        
    #     dt_series = pd.to_datetime(series_str, errors='coerce', format='mixed')
        
    #     # Format valid times, replacing NaT (failed parses) with a safe value like '00:00:00'
    #     time_series = dt_series.dt.strftime('%H:%M:%S').fillna('00:00:00')
        
    #     return time_series

    # def _harmonize_uk_ire_program_descriptions_strict(self) -> Dict[str, Any]:
        """
        Harmonizes Program Description from IRELAND (Source) to UK (Target) for matching 
        time slots, using the 3-minute delta and complex channel matching rules.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_UK_IRE_Harmonization_Flag'
        TIME_TOLERANCE_MINUTES = 3 
        
        # Define Source and Target Markets (Normalized names)
        UK_MARKETS = ['UNITED KINGDOM', 'UK'] # Target of correction
        IRELAND_MARKET = 'IRELAND'            # Source of accurate description

        # Get the static channel classification map
        CHANNEL_MAP = self._get_uk_ire_channel_map()
        
        # 1. Initialization and Checks
        self.df[FLAG_COLUMN] = 'OK'
        REQUIRED_COLS = ['Market', 'Program Description', 'Date', 'Start', 'TV-Channel']
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {
                "check_key": "harmonize_uk_ire_desc_strict", "status": "Skipped",
                "action": "Description Harmonization", 
                "description": "Skipped: Missing required BSR columns.",
                "details": {"rows_updated": 0}
            }

        # --- PART 1: Data Preparation ---
        
        df_temp = self.df.copy()
        
        # 🚨 CRITICAL FIX: Sanitize ALL columns used for filtering or comparison 
        # to guarantee they are strings and safe from the 'float' error.
        TEXT_COLS_TO_SANITIZE = ['Market', 'TV-Channel', 'Program Description', 'Date', 'Start']
        for col in TEXT_COLS_TO_SANITIZE:
            if col in df_temp.columns:
                df_temp[col] = df_temp[col].astype(str).str.strip()

        df_temp['Market_Norm'] = df_temp['Market'].str.upper()
        df_temp['Channel_Norm'] = df_temp['TV-Channel'].str.upper() # Base Channel name used in map lookup

        # Ensure Date/Start are clean datetime objects for calculation
        try:
            # 1. Clean the Start time column using the safe helper
            start_time_clean = self._safe_get_time_string(df_temp['Start'])
            
            # 2. Re-parse the Date 
            date_clean = pd.to_datetime(df_temp['Date'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('1970-01-01')
            
            # 3. Combine valid Date and clean Time
            df_temp['DateTime_Key'] = pd.to_datetime(date_clean + ' ' + start_time_clean, errors='coerce')
            
            if df_temp['DateTime_Key'].isna().sum() > (len(df_temp) * 0.1): 
                raise ValueError("Excessive NaT values after concatenation.")
                
        except Exception as e:
            return {"check_key": "harmonize_uk_ire_desc_strict", "status": "Failed", "action": "Description Harmonization", 
                    "description": f"Failed to create DateTime Key due to parsing errors: {e}", "details": {"rows_updated": 0}}

        # --- PART 2: Match and Update Logic (Focusing on the Comparison) ---

        # 1. Isolate SOURCE Data (Ireland Rows)
        ireland_source_mask = df_temp['Market_Norm'] == IRELAND_MARKET
        ireland_source_df = df_temp[ireland_source_mask].copy()

        # 2. Isolate TARGET Data (UK Rows)
        uk_target_mask = df_temp['Market_Norm'].isin(UK_MARKETS)
        uk_target_df = df_temp[uk_target_mask].copy()

        rows_updated = 0
        
        # Get the static channel classification map
        CHANNEL_MAP = self._get_uk_ire_channel_map()
        
        # Iterate through every clean Ireland broadcast slot
        for ire_index, ire_row in ireland_source_df.iterrows():
            
            ire_dt = ire_row['DateTime_Key']
            ire_channel_norm = ire_row['Channel_Norm']
            ire_desc = ire_row['Program Description'] # This is guaranteed clean string
            
            canonical_ire_id = CHANNEL_MAP.get(ire_channel_norm, ire_channel_norm) 

            # --- A. Time Alignment Check (3-Minute Delta) ---
            same_day_mask = uk_target_df['DateTime_Key'].dt.date == ire_dt.date()
            time_diff = (uk_target_df['DateTime_Key'] - ire_dt).abs() / timedelta(minutes=1)
            time_match_mask = time_diff <= TIME_TOLERANCE_MINUTES
            
            # --- B. Channel Match Check ---
            uk_time_aligned_df = uk_target_df[same_day_mask & time_match_mask].copy()
            
            if not uk_time_aligned_df.empty:
                
                for uk_index, uk_row in uk_time_aligned_df.iterrows():
                    
                    uk_channel_norm = uk_row['Channel_Norm']
                    uk_original_desc = uk_row['Program Description'] # This is guaranteed clean string
                    
                    canonical_uk_id = CHANNEL_MAP.get(uk_channel_norm, uk_channel_norm)
                    
                    # Check 1: Do the two canonical channel IDs match?
                    if canonical_uk_id == canonical_ire_id:
                        
                        # Check 2: Are the descriptions DIFFERENT? (The comparison is now safe)
                        if uk_original_desc != ire_desc:
                            
                            # Action: Overwrite the UK description in the original DF
                            self.df.loc[uk_index, 'Program Description'] = ire_desc
                            
                            # Flag the row
                            self.df.loc[uk_index, FLAG_COLUMN] = f"Description Harmonized from IRELAND: {ire_channel_norm}"
                            rows_updated += 1
                            
        # Final cleanup of temporary columns
        self.df.drop(columns=['DateTime_Key', 'Market_Norm', 'Channel_Norm'] + TEXT_COLS_TO_SANITIZE, errors='ignore', inplace=True)
        
        return {
            "check_key": "harmonize_uk_ire_desc_strict",
            "status": "Flagged" if rows_updated > 0 else "Completed",
            "action": "Description Harmonization (Strict)", 
            "description": f"Harmonized IRELAND descriptions to UK for {rows_updated} matching time slots (3min tolerance).",
            "details": {"rows_updated": int(rows_updated), "time_tolerance_min": TIME_TOLERANCE_MINUTES}
        }

    # --- HELPER 1: Channel List and Mapping (Defines the Restricted Scope) ---

    def _safe_get_time_string(self, series):
        """Safely extracts time string from mixed data types, enforcing string conversion."""
        series_str = series.astype(str) 
        dt_series = pd.to_datetime(series_str, errors='coerce', format='mixed')
        time_series = dt_series.dt.strftime('%H:%M:%S').fillna('00:00:00')
        return time_series

    # --- HELPER 2: Channel List and Mapping (Defines the Restricted Scope) ---

    def _get_target_channel_map(self):
        """
        Creates a standardized map of approved UK/Ireland channels, classifying them 
        by their canonical name for pairing purposes.
        """
        # NOTE: The provided list contains duplicates for "count=2" channels, 
        # indicating they use the same name for both UK and IRE feeds.
        RAW_APPROVED_CHANNELS = [
            'BBC1', 'BBC1', 'BBC2', 'BBC2', 'Channel 4', 'Channel 4', 
            'Premier Sports 1 IRE', 'Premier Sports 1 UK', 'Quest', 
            'Sky Mix', 'Sky Mix', 'Sky News', 'Sky News UK', 
            'Sky Showcase IRE', 'Sky Showcase UK', 'Sky Sports Football', 
            'Sky Sports Football', 'Sky Sports Golf', 'Sky Sports Golf UK',
            'Sky Sports Main Event', 'Sky Sports Main Event', 'Sky Sports News', 
            'Sky Sports News UK', 'Sky Sports Premier League', 'Sky Sports Premier League UK', 
            'Sky Sports Racing', 'Sky Sports Racing', 'Sky Sports Tennis IRE', 
            'Sky Sports Tennis UK', 'TNT Sports 1', 'TNT Sports 1', 'U+Dave', 'U+Dave'
        ]

        channel_map = {}
        
        for raw_name in RAW_APPROVED_CHANNELS:
            # Standardize the base name (remove suffixes for core identity)
            base_name = raw_name.replace(' IRE', '').replace(' UK', '').strip().upper()
            
            # We classify the channel by its BASE name.
            if base_name not in channel_map:
                channel_map[base_name] = base_name
                
        # The map only contains the unique base channel names: {'BBC1': 'BBC1', 'SKY NEWS': 'SKY NEWS', ...}
        return channel_map

    # --- CORE HARMONIZATION FUNCTION (Final Logic) ---

    def _harmonize_uk_ire_program_descriptions_simple(self) -> Dict[str, Any]:
        """
        Audits Program Description consistency between IRELAND (Source) and UK (Target).
        Flags BOTH rows if a time-aligned, channel-matched pair has different descriptions, 
        but DOES NOT perform the descriptive overwrite.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_UK_IRE_Harmonization_Flag'
        TIME_TOLERANCE_MINUTES = 3 
        
        UK_MARKETS = ['UNITED KINGDOM', 'UK']
        IRELAND_MARKET = 'IRELAND'

        self.df[FLAG_COLUMN] = 'OK'
        REQUIRED_COLS = ['Market', 'Program Description', 'Date', 'Start', 'TV-Channel']
        if not all(col in self.df.columns for col in REQUIRED_COLS):
            return {"check_key": "harmonize_uk_ire_desc_simple", "status": "Skipped", "action": "Description Harmonization", "description": "Skipped: Missing required BSR columns.", "details": {"rows_flagged": 0}}

        # Get the strict map of approved channel base names
        APPROVED_CHANNEL_BASES = set(self._get_target_channel_map().keys()) # Assumed helper call

        # --- PART 1: Data Preparation ---
        df_temp = self.df.copy()
        
        # Normalize Market and Channel names
        df_temp['Market_Norm'] = df_temp['Market'].astype(str).str.upper().str.strip()
        df_temp['Canonical_Channel'] = self._create_canonical_channel_key(df_temp['TV-Channel'])

        # Time parsing (assumed functional helper call)
        try:
            start_time_clean = self._safe_get_time_string(df_temp['Start'])
            date_clean = pd.to_datetime(df_temp['Date'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('1970-01-01')
            df_temp['DateTime_Key'] = pd.to_datetime(date_clean + ' ' + start_time_clean, errors='coerce')
        except Exception:
            return {"check_key": "harmonize_uk_ire_desc_simple", "status": "Failed", "action": "Description Harmonization", "description": "Failed to create DateTime Key.", "details": {"rows_flagged": 0}}

        # --- PART 2: Match and Audit Logic ---

        # 1. Isolate SOURCE Data (Ireland Rows)
        ireland_source_mask = df_temp['Market_Norm'] == IRELAND_MARKET
        ireland_source_mask = ireland_source_mask & df_temp['Canonical_Channel'].isin(APPROVED_CHANNEL_BASES)
        ireland_source_df = df_temp[ireland_source_mask].copy()

        # 2. Isolate TARGET Data (UK Rows)
        uk_target_mask = df_temp['Market_Norm'].isin(UK_MARKETS)
        uk_target_mask = uk_target_mask & df_temp['Canonical_Channel'].isin(APPROVED_CHANNEL_BASES)
        uk_target_df = df_temp[uk_target_mask].copy()

        rows_flagged = 0
        flagged_indices = set() # Collects indices from both UK and IRE involved in a mismatch
        
        # Iterate through every clean Ireland broadcast slot (Source)
        for ire_index, ire_row in ireland_source_df.iterrows():
            
            ire_dt = ire_row['DateTime_Key']
            ire_key = ire_row['Canonical_Channel']
            ire_desc = ire_row['Program Description']
            
            # --- Find UK Matches ---
            same_day_mask = uk_target_df['DateTime_Key'].dt.date == ire_dt.date()
            time_diff = (uk_target_df['DateTime_Key'] - ire_dt).abs() / timedelta(minutes=1)
            time_match_mask = time_diff <= TIME_TOLERANCE_MINUTES
            channel_match_mask = uk_target_df['Canonical_Channel'] == ire_key
            final_match_mask = same_day_mask & time_match_mask & channel_match_mask
            
            uk_matching_rows = uk_target_df[final_match_mask]
            
            if not uk_matching_rows.empty:
                
                for uk_index, uk_row in uk_matching_rows.iterrows():
                    uk_original_desc = uk_row['Program Description']
                    
                    # Check: Are the descriptions DIFFERENT? (This is the anomaly)
                    if uk_original_desc.strip() != ire_desc.strip():
                        
                        # Flagging Action: Record both source (IRE) and target (UK) indices
                        flagged_indices.add(ire_index) # Flag the Ireland source row
                        flagged_indices.add(uk_index)  # Flag the UK target row

        # --- 3. Apply Flag to Original DataFrame ---
        rows_flagged = len(flagged_indices)
        
        if rows_flagged > 0:
            flag_message = f"DESCRIPTION MISMATCH: Descriptions differ on time-aligned channel slot (Review UK/IRE data)."
            
            # Apply the flag to all identified indices (both UK and IRE)
            self.df.loc[list(flagged_indices), FLAG_COLUMN] = flag_message
                            
        # Final cleanup of temporary columns
        self.df.drop(columns=['DateTime_Key', 'Market_Norm', 'Channel_Norm', 'Canonical_Channel'], errors='ignore', inplace=True)
        
        return {
            "check_key": "harmonize_uk_ire_desc_simple",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Description Consistency Audit", 
            "description": f"Audited description consistency. Flagged {rows_flagged} rows in pairs where UK and IRELAND descriptions differed.",
            "details": {
                "rows_flagged": int(rows_flagged),
                "time_tolerance_min": TIME_TOLERANCE_MINUTES,
            }
        }

    def _check_game_of_the_day_match(self) -> Dict[str, Any]:
        """
        Identifies BSR rows marked as 'Game of the day' in the UK and checks if 
        corresponding LIVE data exists in the OVN/CDT file based on DATE and TIME only.
        
        CRITERIA:
        1. BSR: 'Game of the day', UK Market, Non-empty Teams.
        2. OVN: 'Live' Type of programme.
        3. Match: Same Date, Start Time within 5 minutes (Channel ignored).
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Game_of_Day_OVN_Flag'
        KEYWORD = "GAME OF THE DAY"
        TARGET_MARKET_REGEX = r'UNITED KINGDOM|UK'
        TIME_TOLERANCE_MINUTES = 5
        
        self.df[FLAG_COLUMN] = 'OK'
        
        REQUIRED_BSR_COLS = ['Program Description', 'Date (UTC/GMT)', 'Start (UTC)', 'Market', 'Home Team', 'Away Team']
        if not all(col in self.df.columns for col in REQUIRED_BSR_COLS):
            return {
                "check_key": "check_game_of_the_day_match", "status": "Skipped",
                "action": "Game of Day Audit", 
                "description": "Skipped: Missing required BSR columns.",
                "details": {"rows_flagged": 0}
            }

        # 1. Load OVN Data
        # Ensure your _load_overnight_data function returns 'Type of programme'
        # If it doesn't, you may need to update the loader columns list.
        df_ovn = self._load_overnight_data()
        
        if df_ovn is None or df_ovn.empty:
            return {"check_key": "check_game_of_the_day_match", "status": "Skipped", "description": "Skipped: Overnight file not loaded."}

        # --- Filter OVN for 'Live' ---
        # Normalize column name check just in case
        type_col = next((c for c in df_ovn.columns if 'Type of programme' in c), None)
        
        if type_col:
            # Filter for Live/Near Live
            df_ovn_live = df_ovn[df_ovn[type_col].astype(str).str.lower().str.contains('live', na=False)].copy()
        else:
            # Fallback: If column missing, assume all rows are valid candidates (or log warning)
            df_ovn_live = df_ovn.copy()

        # 2. Filter BSR Targets
        market_mask = self.df['Market'].astype(str).str.upper().str.contains(TARGET_MARKET_REGEX, regex=True, na=False)
        desc_mask = self.df['Program Description'].astype(str).str.upper().str.contains(KEYWORD, na=False)
        teams_mask = (self.df['Home Team'].notna()) & (self.df['Home Team'].astype(str).str.strip() != '') & \
                    (self.df['Away Team'].notna()) & (self.df['Away Team'].astype(str).str.strip() != '')
        
        target_bsr_mask = market_mask & desc_mask & teams_mask
        
        if not target_bsr_mask.any():
            return {"check_key": "check_game_of_the_day_match", "status": "Completed", "action": "Game of Day Audit", "description": "No valid 'Game of the day' targets found.", "details": {"rows_flagged": 0}}

        # 3. Prepare Data for Merge
        try:
            # BSR Prep
            df_bsr_targets = self.df.loc[target_bsr_mask].copy()
            df_bsr_targets['BSR_Index'] = df_bsr_targets.index
            
            # Clean Dates (Common Key)
            df_bsr_targets['Date_Join'] = pd.to_datetime(df_bsr_targets['Date (UTC/GMT)'], errors='coerce').dt.strftime('%Y-%m-%d')
            df_ovn_live['Date_Join'] = pd.to_datetime(df_ovn_live['Date'], errors='coerce').dt.strftime('%Y-%m-%d')

            # Clean Times (For Delta Check)
            # BSR
            start_str_bsr = self._safe_get_time_string(df_bsr_targets['Start (UTC)'])
            df_bsr_targets['Start_DT'] = pd.to_datetime(df_bsr_targets['Date_Join'] + ' ' + start_str_bsr)
            
            # OVN
            start_str_ovn = self._safe_get_time_string(df_ovn_live['Start Time'])
            df_ovn_live['Start_DT'] = pd.to_datetime(df_ovn_live['Date_Join'] + ' ' + start_str_ovn)

            # 4. Perform Match (Join on Date, Filter on Time)
            # Inner join creates a Cartesian product for that day (comparing every BSR row to every OVN row on that date)
            merged = pd.merge(
                df_bsr_targets,
                df_ovn_live[['Date_Join', 'Start_DT']], # Only bring necessary cols
                on='Date_Join',
                suffixes=('_BSR', '_OVN')
            )
            
            # Calculate Absolute Time Difference
            merged['Time_Diff'] = (merged['Start_DT_BSR'] - merged['Start_DT_OVN']).abs()
            
            # Filter for matches within tolerance
            matches = merged[merged['Time_Diff'] <= timedelta(minutes=TIME_TOLERANCE_MINUTES)]
            
            # Get unique BSR indices that found a match
            indices_to_flag = matches['BSR_Index'].unique()
            rows_flagged = len(indices_to_flag)
            
            # 5. Apply Flag
            if rows_flagged > 0:
                flag_msg = "UPDATE REQUIRED: OVN Live Match Available (Date Match + Time <= 5min)."
                self.df.loc[indices_to_flag, FLAG_COLUMN] = flag_msg

        except Exception as e:
            return {"check_key": "check_game_of_the_day_match", "status": "Failed", "description": f"Error during matching: {e}"}

        return {
            "check_key": "check_game_of_the_day_match",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Game of Day Audit", 
            "description": f"Flagged {rows_flagged} 'Game of the day' rows with matching Live OVN data.",
            "details": {
                "rows_flagged": int(rows_flagged),
                "ovn_records_checked": len(df_ovn_live)
            }
        }

    # def _check_non_metered_primary_market_audience(self) -> Dict[str, Any]:
        """
        Audits the 'Source' column to identify rows that are duplicated from a 
        non-metered primary market (Source = 'BSA' only, without 'Time bans' or 'BC-LOGS').
        Sets audience values to 0 for these rows.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Non_Metered_Audience_Flag'
        
        SOURCE_COL = 'Source'
        # Target columns to zero out
        AUDIENCE_COLS = ['Aud. Estimates [\'000s]', 'Aud Metered (000s) 3+']
        
        # Keywords
        KEYWORD_BSA = 'BSA'
        KEYWORDS_VALIDATORS = ['TIME BANS', 'BC-LOGS', 'BC LOGS'] # Variations of valid markers
        
        self.df[FLAG_COLUMN] = 'OK'
        
        # Check required columns
        cols_to_check = [SOURCE_COL] + AUDIENCE_COLS
        if not all(col in self.df.columns for col in cols_to_check):
            return {
                "check_key": "check_non_metered_audience", "status": "Skipped",
                "action": "Non-Metered Audience Check", 
                "description": "Skipped: Missing Source or Audience columns.",
                "details": {"rows_suppressed": 0}
            }

        # 1. Normalize Source Column
        source_norm = self.df[SOURCE_COL].astype(str).str.upper().str.strip()
        
        # 2. Create Masks
        
        # Condition A: Contains "BSA"
        has_bsa = source_norm.str.contains(KEYWORD_BSA, na=False)
        
        # Condition B: Contains "Time bans" OR "BC-LOGS"
        # Create regex pattern: "TIME BANS|BC-LOGS|BC LOGS"
        validator_pattern = '|'.join([re.escape(k) for k in KEYWORDS_VALIDATORS])
        has_validator = source_norm.str.contains(validator_pattern, regex=True, na=False)
        
        # 3. Identify Rows to Suppress
        # Logic: Has BSA AND DOES NOT HAVE Validator
        suppression_mask = has_bsa & (~has_validator)
        
        rows_suppressed = suppression_mask.sum()
        
        # 4. Apply Update and Flag
        if rows_suppressed > 0:
            
            # Action 1: Set Audience columns to 0
            for col in AUDIENCE_COLS:
                if col in self.df.columns:
                    self.df.loc[suppression_mask, col] = 0.0
            
            # Action 2: Flag the row
            self.df.loc[suppression_mask, FLAG_COLUMN] = "AUDIENCE SUPPRESSED: Source is BSA only (Non-Metered Primary Market)."

        return {
            "check_key": "check_non_metered_audience",
            "status": "Flagged" if rows_suppressed > 0 else "Completed",
            "action": "Non-Metered Audience Check", 
            "description": f"Suppressed audience for {rows_suppressed} rows where Source indicated non-metered origin (BSA without validation).",
            "details": {
                "rows_suppressed": int(rows_suppressed),
                "bsr_source_criteria": "BSA only (No Time bans/BC-LOGS)"
            }
        }
    def _check_non_metered_primary_market_audience(self) -> Dict[str, Any]:
        """
        Audits the 'Source' column to identify rows that come from a non-metered 
        primary market (Source = 'BSA' only). 
        Flags the row if Audience data is present when it should be zero.
        Does NOT modify the audience data.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Non_Metered_Audience_Flag'
        
        SOURCE_COL = 'Source'
        AUDIENCE_COLS = ['Aud. Estimates [\'000s]', 'Aud Metered (000s) 3+']
        
        KEYWORD_BSA = 'BSA'
        KEYWORDS_VALIDATORS = ['TIME BANS', 'BC-LOGS', 'BC LOGS']
        
        self.df[FLAG_COLUMN] = 'OK'
        
        # Check required columns
        cols_to_check = [SOURCE_COL] + [col for col in AUDIENCE_COLS if col in self.df.columns]
        if not all(col in self.df.columns for col in cols_to_check):
            return {
                "check_key": "check_non_metered_audience", "status": "Skipped",
                "action": "Non-Metered Audience Audit", 
                "description": "Skipped: Missing Source or Audience columns.",
                "details": {"rows_flagged": 0}
            }

        # 1. Normalize Source Column
        source_norm = self.df[SOURCE_COL].astype(str).str.upper().str.strip()
        
        # 2. Create Source Masks
        has_bsa = source_norm.str.contains(KEYWORD_BSA, na=False)
        
        validator_pattern = '|'.join([re.escape(k) for k in KEYWORDS_VALIDATORS])
        has_validator = source_norm.str.contains(validator_pattern, regex=True, na=False)
        
        # Mask A: Logic for Non-Metered Source (Has BSA but NO Validator)
        is_non_metered_source = has_bsa & (~has_validator)
        
        # 3. Create Audience Mask (Check if data actually exists > 0)
        # We only want to flag if there is a value to suppress. If it's already 0, it's fine.
        audience_df = self.df[cols_to_check].drop(columns=[SOURCE_COL])
        # Coerce to numeric and check if any audience column has a value > 0
        has_data_to_remove = (audience_df.apply(pd.to_numeric, errors='coerce').fillna(0) > 0).any(axis=1)
        
        # 4. Combine Masks
        flag_mask = is_non_metered_source & has_data_to_remove
        
        rows_flagged = flag_mask.sum()
        
        # 5. Apply Flag (NO DATA CHANGE)
        if rows_flagged > 0:
            flag_message = "AUDIENCE ANOMALY: Source is BSA-only (Non-Metered), but Audience > 0. Should be suppressed."
            
            # Apply flag only to rows currently marked OK to avoid overwriting higher priority errors
            rows_to_flag = flag_mask & (self.df[FLAG_COLUMN] == 'OK')
            self.df.loc[rows_to_flag, FLAG_COLUMN] = flag_message

        return {
            "check_key": "check_non_metered_audience",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Non-Metered Audience Audit", 
            "description": f"Flagged {rows_flagged} rows where Source indicates non-metered origin but Audience data is present.",
            "details": {
                "rows_flagged": int(rows_flagged),
                "bsr_source_criteria": "BSA only (No Time bans/BC-LOGS)"
            }
        }

    def _check_legacy_mapping(self) -> Dict[str, Any]:
        """
        Checks BSR rows against a Legacy Mapping sheet. 
        - If BSR matches 'Original' and differs from 'Legacy': Flags "Legacy Definition Found".
        - If BSR matches 'Original' and matches 'Legacy': Flags "Both are Same".
        - If BSR row is not in the Legacy Mapping: Flags "Not found in the obligation sheet".
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_Legacy_Mapping_Flag'
        
        self.df[FLAG_COLUMN] = 'OK'
        
        # 1. Load the Multi-Index DataFrame (Legacy Mapping)
        # NOTE: Assumes _load_full_obligation_data is the function loading the Legacy sheet (header=[0,1])
        df_map = self._load_full_obligation_data()
        
        if df_map is None or df_map.empty:
            return {
                "check_key": "check_legacy_mapping", "status": "Failed",
                "action": "Legacy Mapping Check", 
                "description": "Error: Legacy Mapping sheet needed but not found or empty.",
                "details": {"rows_flagged": 0}
            }

        # 2. Prepare Matching Columns & Comparison Logic
        try:
            # --- ORIGINAL COLUMNS (The Key) ---
            map_orig_market = df_map[('Original details', 'Market')].astype(str).str.strip().str.upper()
            map_orig_channel = df_map[('Original details', 'TV-Channel')].astype(str).str.strip().str.upper()
            
            # --- LEGACY COLUMNS (The Value) ---
            map_leg_market_raw = df_map[('Legacy details', 'Market')].astype(str).str.strip()
            map_leg_channel_raw = df_map[('Legacy details', 'TV-Channel')].astype(str).str.strip()
            map_leg_id = df_map[('Legacy details', 'Channel ID')]
            
            # Create normalized versions for comparison (Upper case)
            map_leg_market_norm = map_leg_market_raw.str.upper()
            map_leg_channel_norm = map_leg_channel_raw.str.upper()
            
            # Create a temporary mapping DataFrame
            flat_map = pd.DataFrame({
                'Match_Key': map_orig_market + '|' + map_orig_channel,
                'Legacy_Market_Norm': map_leg_market_norm,
                'Legacy_Channel_Norm': map_leg_channel_norm,
                'Legacy_Display_Info': map_leg_market_raw + ' | ' + map_leg_channel_raw + ' (ID: ' + map_leg_id.astype(str) + ')'
            })
            
        except KeyError as e:
            return {
                "check_key": "check_legacy_mapping", "status": "Failed",
                "action": "Legacy Mapping Check", 
                "description": f"Failed to parse Legacy headers. Structure mismatch: {e}",
                "details": {"rows_flagged": 0}
            }

        # 3. Prepare BSR Keys
        bsr_market_norm = self.df['Market'].astype(str).str.strip().str.upper()
        bsr_channel_norm = self.df['TV-Channel'].astype(str).str.strip().str.upper()
        
        # Create keys for BSR
        self.df['Temp_Legacy_Key'] = bsr_market_norm + '|' + bsr_channel_norm
        
        # 4. Perform Lookups
        # We need to map three things: The Info String, The Legacy Market (Norm), The Legacy Channel (Norm)
        info_dict = flat_map.set_index('Match_Key')['Legacy_Display_Info'].to_dict()
        market_dict = flat_map.set_index('Match_Key')['Legacy_Market_Norm'].to_dict()
        channel_dict = flat_map.set_index('Match_Key')['Legacy_Channel_Norm'].to_dict()
        
        mapped_info = self.df['Temp_Legacy_Key'].map(info_dict)
        mapped_market = self.df['Temp_Legacy_Key'].map(market_dict)
        mapped_channel = self.df['Temp_Legacy_Key'].map(channel_dict)
        
        # 5. Apply Logic
        
        # Mask: Row was found in the mapping
        mask_found = matches = mapped_info.notna()
        mask_not_found = mapped_info.isna()
        
        rows_flagged = 0
        rows_different = 0
        rows_same = 0

        # Handle Not Found Rows
        if mask_not_found.any():
            self.df.loc[mask_not_found, FLAG_COLUMN] = "Not found in the obligation sheet"
            rows_flagged += mask_not_found.sum()

        # Handle Found Rows
        if mask_found.any():
            # Mask: BSR Data matches Legacy Data (Same Market AND Same Channel)
            is_same_mask = mask_found & (bsr_market_norm == mapped_market) & (bsr_channel_norm == mapped_channel)
            
            # Mask: BSR Data differs from Legacy Data (Needs Update)
            is_diff_mask = mask_found & (~is_same_mask)
            
            # Case 1: Difference Found
            if is_diff_mask.any():
                self.df.loc[is_diff_mask, FLAG_COLUMN] = "Legacy Definition Found: Should be " + mapped_info[is_diff_mask]
                rows_different = is_diff_mask.sum()
                rows_flagged += rows_different
            
            # Case 2: Both are Same
            if is_same_mask.any():
                self.df.loc[is_same_mask, FLAG_COLUMN] = "Both are Same"
                rows_same = is_same_mask.sum()

        # Cleanup
        self.df.drop(columns=['Temp_Legacy_Key'], inplace=True, errors='ignore')

        return {
            "check_key": "check_legacy_mapping",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "Legacy Mapping Check", 
            "description": f"Checked against Legacy Map. Flagged {rows_flagged} rows (Diffs/Not Found).",
            "details": {
                "rows_flagged": int(rows_flagged),
                "rows_not_found": int(mask_not_found.sum()),
                "rows_different": int(rows_different),
                "rows_same": int(rows_same),
                "mappings_loaded": len(flat_map)
            }
        }

    def normalize_channel_name(self, channel_series):
        """
        Removes regional codes, parentheses, suffixes, and numbers to compare channels 
        by their core brand identity (e.g., ESPN, Sky).
        """
        # 1. Ensure string and convert to uppercase
        normalized = channel_series.astype(str).str.strip().str.upper()
        
        # 2. Remove anything inside parentheses (e.g., (ARG), (BOL))
        normalized = normalized.str.replace(r'\s*\([^)]*\)', '', regex=True)
        
        # 3. Remove country codes/acronyms
        normalized = normalized.str.replace(
            r'(\s+ARG|\s+BOL|\s+CHL|\s+PER|\s+SWE|\s+DE|\s+AFR|\s+PCA|\s+COL|\s+ECU|\s+URY|\s+MEX|\s+JPN|\s+LTU|\s+CHE|\s+FRA)', 
            '', 
            flags=re.IGNORECASE, 
            regex=True
        )
        
        # 4. Remove 'SPORT' and 'TV'
        normalized = normalized.str.replace(r'\s+SPORT[S]*', '', regex=True)
        normalized = normalized.str.replace(r'\s+TV', '', regex=True)
        
        # 5. Final cleanup
        normalized = normalized.str.replace(r'\s{2,}', ' ', regex=True).str.strip()
        return normalized

    def _safe_get_time_string(self, series):
        """Safely extracts time string from mixed data types."""
        series_str = series.astype(str) 
        dt_series = pd.to_datetime(series_str, errors='coerce', format='mixed')
        time_series = dt_series.dt.strftime('%H:%M:%S').fillna('00:00:00')
        return time_series

    # --- THE CHECK FUNCTION ---

    def _check_premier_league_october_obligation(self) -> Dict[str, Any]:
        """
        Checks BSR rows against the Obligation sheet (CDT Audiences tab) for specific 
        Premier League matches in October. 
        
        MATCHING LOGIC: Matches on [Channel + Broadcaster] ONLY.
        Ignores Date and Time to prevent mismatches due to timezone/formatting differences.
        """
        initial_rows = len(self.df)
        FLAG_COLUMN = 'QC_PL_Oct_Obligation_Match_Flag'
        
        # Constants for filtering
        TARGET_COMPETITION = 'Premier League'
        TARGET_MONTH = 10 # October
        CDT_SHEET_NAME = "CDT Audiences"
        
        self.df[FLAG_COLUMN] = 'OK'

        if not self.obligation_path:
            return {
                "check_key": "check_pl_oct_obligation", "status": "Skipped",
                "action": "PL October Obligation Check", 
                "description": "Skipped: Obligation file not provided.",
                "details": {"rows_flagged": 0}
            }
            
        try:
            # 1. Load CDT sheet (Header row 9 -> Index 8)
            df_obl = pd.read_excel(self.obligation_path, sheet_name=CDT_SHEET_NAME, header=8) 
            df_obl.columns = [str(c).strip() for c in df_obl.columns]
            
            # Verify required columns exist
            if 'Competition' not in df_obl.columns:
                 return {"check_key": "check_pl_oct_obligation", "status": "Failed", "description": f"Column 'Competition' not found in Obligation file. Loaded: {df_obl.columns.tolist()}"}

            # Filter for Premier League
            pl_mask = df_obl['Competition'].astype(str).str.contains(TARGET_COMPETITION, case=False, na=False)
            df_obl_pl = df_obl[pl_mask].copy()

            # Filter for October Dates (This ensures we only look at the relevant month in the OVN file)
            if 'Date' in df_obl_pl.columns:
                df_obl_pl['Date_dt'] = pd.to_datetime(df_obl_pl['Date'], errors='coerce')
                oct_mask = df_obl_pl['Date_dt'].dt.month == TARGET_MONTH
                df_obl_final = df_obl_pl[oct_mask].copy()
            else:
                return {"check_key": "check_pl_oct_obligation", "status": "Failed", "description": "Column 'Date' not found in Obligation file."}
                
            if df_obl_final.empty:
                return {"check_key": "check_pl_oct_obligation", "status": "Completed", "description": "No Premier League obligation data found for October."}

            # 2. Prepare Keys for Matching (Obligation Side)
            # Normalize for matching
            df_obl_final['Channel_Norm'] = self.normalize_channel_name(df_obl_final['Channel'])
            df_obl_final['Broadcaster_Norm'] = df_obl_final['Broadcaster'].astype(str).str.strip().str.upper()
            
            # --- MODIFIED KEY: Removed Date and Start Time ---
            # Matching only on Broadcaster + Channel
            df_obl_final['Match_Key'] = (
                df_obl_final['Channel_Norm'] + '|' + 
                df_obl_final['Broadcaster_Norm']
            )
            
            required_keys = set(df_obl_final['Match_Key'].unique())

        except Exception as e:
            return {"check_key": "check_pl_oct_obligation", "status": "Failed", "description": f"Error processing Obligation data: {e}"}

        # 3. Prepare BSR Data for Matching
        try:
            bsr_check = self.df.copy()
            bsr_check['Channel_Norm'] = self.normalize_channel_name(bsr_check['TV-Channel'])
            bsr_check['Broadcaster_Norm'] = bsr_check['Broadcaster'].astype(str).str.strip().str.upper()
            
            # --- MODIFIED KEY: Removed Date and Start Time ---
            bsr_check['Match_Key'] = (
                bsr_check['Channel_Norm'] + '|' + 
                bsr_check['Broadcaster_Norm']
            )
            
            # 4. Find Matches
            match_mask = bsr_check['Match_Key'].isin(required_keys)
            indices_matched = bsr_check[match_mask].index
            
            rows_flagged = len(indices_matched)
            
            if rows_flagged > 0:
                flag_msg = "Obligation Match Found: Corresponding PL/Oct data exists in Obligation Sheet."
                self.df.loc[indices_matched, FLAG_COLUMN] = flag_msg

        except Exception as e:
            return {"check_key": "check_pl_oct_obligation", "status": "Failed", "description": f"Error processing BSR data for match: {e}"}

        return {
            "check_key": "check_pl_oct_obligation",
            "status": "Flagged" if rows_flagged > 0 else "Completed",
            "action": "PL October Obligation Check", 
            "description": f"Flagged {rows_flagged} BSR rows matching Premier League/October obligation channels (Ignoring Date/Time).",
            "details": {
                "rows_flagged": int(rows_flagged),
                "obligation_entries_count": len(required_keys)
            }
        }

    
    def _filter_short_programs(self):
        """
        Removes programs where duration <5 minutes except Austria and New Zealand.
        Stores removed rows in: self.short_programs_df
        """
        MIN_DURATION = 5
        EXEMPT = ["AUSTRIA", "NEW ZEALAND"]

        df = self.df.copy()

        # Normalize markets
        df["Market_norm"] = df["Market"].astype(str).str.upper()

        # Parse start + end with date included
        df["Date_only"] = pd.to_datetime(df["Date"], errors="coerce").dt.date.astype(str)

        df["Start_DT"] = pd.to_datetime(
            df["Date_only"] + " " + df["Start (UTC)"].astype(str),
            errors="coerce"
        )
        df["End_DT_raw"] = pd.to_datetime(
            df["Date_only"] + " " + df["End (UTC)"].astype(str),
            errors="coerce"
        )

        # Handle past-midnight rollover
        rollover = df["End_DT_raw"] < df["Start_DT"]
        df.loc[rollover, "End_DT_raw"] += pd.Timedelta(days=1)

        # Compute duration in minutes
        df["Duration_Min"] = (df["End_DT_raw"] - df["Start_DT"]).dt.total_seconds() / 60

        remove_mask = (df["Duration_Min"] < MIN_DURATION) & (~df["Market_norm"].isin(EXEMPT))

        removed_df = df[remove_mask].copy()
        keep_df = df[~remove_mask].copy()

        # Clean temp cols
        for col in ["Market_norm", "Start_DT", "End_DT_raw", "Duration_Min", "Date_only"]:
            removed_df.drop(columns=col, inplace=True, errors="ignore")
            keep_df.drop(columns=col, inplace=True, errors="ignore")

        self.short_programs_df = removed_df
        self.df = keep_df

        return {
            "check_key": "filter_short_programs",
            "status": "Flagged" if len(removed_df) else "Completed",
            "description": f"{len(removed_df)} short programs removed (<5 min)",
            "details": {}
        }
    
    def _sa_nielsen_inclusion_check(self):
        """
        SA Nielsen Inclusion Check (Case Sensitive):
        Extract rows where Market exactly equals "South Africa".
        Remove them from the main DF and store in self.sa_nielsen_df.
        """
        TARGET = "South Africa"

        if "Market" not in self.df.columns:
            return {
                "check_key": "sa_nielsen_inclusion_check",
                "status": "Skipped",
                "description": "Column 'Market' not found.",
                "details": {}
            }

        # Strict case-sensitive match
        sa_mask = self.df["Market"] == TARGET

        sa_rows = self.df.loc[sa_mask].copy()
        remaining = self.df.loc[~sa_mask].copy()

        # Save new versions
        self.sa_nielsen_df = sa_rows.reset_index(drop=True)
        self.df = remaining.reset_index(drop=True)

        return {
            "check_key": "sa_nielsen_inclusion_check",
            "status": "Completed",
            "description": f"Extracted {len(sa_rows)} rows for SA Nielsen tab.",
            "details": {"rows_found": int(len(sa_rows))}
        }
    
    def _epl_live_vs_delay_validation(self):
        """
        NEW EPL Live vs Delay Validation
        Logic:
        For each broadcast row:
            - Match fixture using home, away, date
            - Compare Start(UTC) with fixture kickoff time
            - If within ±60 minutes -> LIVE
            - Else -> DELAYED
        Output:
        - Adds 'EPL_LiveDelay_Flag'
        - Produces self.live_delay_flags_df
        """

        FLAG = "EPL_LiveDelay_Flag"
        self.df[FLAG] = ""

        # ----------------------------
        # Load fixture sheet 
        # ----------------------------
        bsr_path = self.bsr_path
        try:
            xl = pd.ExcelFile(bsr_path)
        except:
            return {
                "check_key": "epl_live_vs_delay_validation",
                "status": "Skipped",
                "description": "Unable to read BSR file for fixture sheet.",
                "details": {}
            }

        # detect fixture sheet
        fixture_keywords = ["fixture", "fixtures"]
        fixture_sheet = None
        for s in xl.sheet_names:
            if any(k in s.lower() for k in fixture_keywords):
                fixture_sheet = s
                break

        if fixture_sheet is None:
            return {
                "check_key": "epl_live_vs_delay_validation",
                "status": "Skipped",
                "description": "Fixture sheet not found",
                "details": {}
            }

        df_fix = xl.parse(fixture_sheet)

        # required columns in fixture
        needed_fix_cols = ["Home Team", "Away Team", "Date", "Start Time", "End Time"]
        for c in needed_fix_cols:
            if c not in df_fix.columns:
                return {
                    "check_key": "epl_live_vs_delay_validation",
                    "status": "Skipped",
                    "description": f"Missing fixture column: {c}",
                    "details": {}
                }

        # ----------------------------
        # Prepare fixture lookup
        # ----------------------------
        def clean(x):
            if pd.isna(x): return ""
            x = str(x).lower().strip()
            x = re.sub(r"[^\w\s]", " ", x)
            return re.sub(r"\s+", " ", x).strip()

        df_fix["_home"] = df_fix["Home Team"].map(clean)
        df_fix["_away"] = df_fix["Away Team"].map(clean)
        df_fix["_date"] = pd.to_datetime(df_fix["Date"], errors="coerce").dt.date

        # Combine fixture datetime
        df_fix["_fix_start"] = [
            pd.to_datetime(str(df_fix.at[i, "Date"]) + " " + str(df_fix.at[i, "Start Time"]), errors="coerce")
            for i in df_fix.index
        ]

        # ----------------------------
        # Prepare BSR
        # ----------------------------
        df = self.df

        required_bsr_cols = ["Home Team", "Away Team", "Date", "Start (UTC)"]
        missing = [c for c in required_bsr_cols if c not in df.columns]
        if missing:
            return {
                "check_key": "epl_live_vs_delay_validation",
                "status": "Skipped",
                "description": f"Missing BSR columns: {missing}",
                "details": {}
            }

        df["_home"] = df["Home Team"].map(clean)
        df["_away"] = df["Away Team"].map(clean)
        df["_date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        df["_bsr_start"] = [
            pd.to_datetime(str(df.at[i, "Date"]) + " " + str(df.at[i, "Start (UTC)"]), errors="coerce")
            for i in df.index
        ]

        # ----------------------------
        # MAIN LOGIC
        # ----------------------------
        tolerance = 60  # minutes
        flagged_rows = []

        for i, row in df.iterrows():
            h = row["_home"]
            a = row["_away"]
            d = row["_date"]
            bsr_start = row["_bsr_start"]

            if pd.isna(bsr_start) or pd.isna(d):
                continue

            # Fixture match
            fx = df_fix[
                (df_fix["_home"] == h)
                & (df_fix["_away"] == a)
                & (df_fix["_date"] == d)
            ]

            if fx.empty:
                # no fixture → skip silently
                continue

            fix_start = fx["_fix_start"].iloc[0]
            if pd.isna(fix_start):
                continue

            diff = abs((bsr_start - fix_start).total_seconds() / 60)

            # LIVE
            if diff <= tolerance:
                df.at[i, FLAG] = f"LIVE: within ±{tolerance} min"
                flagged_rows.append(i)
            else:
                df.at[i, FLAG] = f"DELAYED: diff={diff:.1f} min"
                flagged_rows.append(i)

        # ----------------------------
        # Save flagged rows for export
        # ----------------------------
        self.live_delay_flags_df = df.loc[flagged_rows].copy()

        # Cleanup
        df.drop(columns=["_home", "_away", "_date", "_bsr_start"], errors="ignore", inplace=True)

        total_flagged = len(flagged_rows)
        return {
            "check_key": "epl_live_vs_delay_validation",
            "status": "Flagged" if total_flagged else "Completed",
            "description": f"Checked Live vs Delay using fixture ±1 hour tolerance. Flagged {total_flagged} rows.",
            "details": {"rows_flagged": total_flagged}
        }


    def _pl_magazine_highlights_classification(self):
        """
        EPL: PL Magazine / Highlights Classification

        Auto-detects:
        - 'Type of program' column (any casing/spacing)
        - 'Combined' column

        Works even if names differ.
        """

        df = self.df.copy()

        # ---------- Helper to find column robustly ----------
        def find_col(df, name):
            name = name.strip().lower()
            for col in df.columns:
                if col.strip().lower() == name:
                    return col
            return None

        # Detect required columns
        col_progtype = find_col(df, "type of program")
        col_combined = find_col(df, "combined")

        missing = []
        if col_progtype is None: missing.append("Type of program")
        if col_combined is None: missing.append("Combined")

        if missing:
            return {
                "check_key": "pl_magazine_highlights_classification",
                "status": "Skipped",
                "description": f"Missing required columns: {missing}",
                "details": {}
            }

        # Output column
        CATEGORY_COL = "PL_Magazine_Highlights_Category"
        df[CATEGORY_COL] = ""

        # Normalize
        type_norm = df[col_progtype].astype(str).str.lower().str.strip()
        combined_norm = df[col_combined].astype(str).str.lower()

        # -----------------------------------------------------
        # 1️⃣  MAGAZINE & SUPPORT
        # -----------------------------------------------------
        mag_mask = type_norm == "magazine & support"

        df.loc[mag_mask & combined_norm.str.contains("vault"), CATEGORY_COL] = "PL Magazine"
        df.loc[mag_mask & combined_norm.str.contains("stories"), CATEGORY_COL] = "PL Stories"
        df.loc[mag_mask & combined_norm.str.contains("preview"), CATEGORY_COL] = "PL Preview"
        df.loc[mag_mask & combined_norm.str.contains("the big interview"), CATEGORY_COL] = "PL The Big Interview"

        df.loc[mag_mask & (df[CATEGORY_COL] == ""), CATEGORY_COL] = "PL Magazine"

        # -----------------------------------------------------
        # 2️⃣  HIGHLIGHTS
        # -----------------------------------------------------
        high_mask = type_norm == "highlights"

        df.loc[high_mask & combined_norm.str.contains("netbusters"), CATEGORY_COL] = "PL Netbusters"
        df.loc[high_mask & combined_norm.str.contains("reload"), CATEGORY_COL] = "PL Reload"
        df.loc[high_mask & combined_norm.str.contains("review"), CATEGORY_COL] = "PL Review"
        df.loc[high_mask & combined_norm.str.contains("match of the day"), CATEGORY_COL] = "PL Match Of The Day"

        df.loc[high_mask & (df[CATEGORY_COL] == ""), CATEGORY_COL] = "PL Highlights"

        # -----------------------------------------------------
        # Report tab
        # -----------------------------------------------------
        report_df = df[df[CATEGORY_COL] != ""].copy()
        self.pl_mag_highlights_df = report_df

        # Save back to main DF
        self.df = df

        return {
            "check_key": "pl_magazine_highlights_classification",
            "status": "Completed",
            "description": f"Classified {len(report_df)} rows.",
            "details": {"rows_classified": len(report_df)}
        }


# ----------------------------- ⚙️ Utility Functions (kept standalone) -----------------------------



    def color_excel(output_path, df):
        """Applies green/red coloring based on QC_OK columns."""
        wb = load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_map = {name: idx+1 for idx, name in enumerate(headers)}

        qc_columns = [col for col in df.columns if col.endswith("_OK")]

        for col_name in qc_columns:
            if col_name in col_map:
                col_idx = col_map[col_name]
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    val = cell.value
                    if val in [True, "True"]:
                        cell.fill = GREEN_FILL
                    elif val in [False, "False"]:
                        cell.fill = RED_FILL

        wb.save(output_path)


    def generate_summary_sheet(output_path, df):
        """Generates a summary sheet with pass/fail counts for QC checks."""
        wb = load_workbook(output_path)
        if "Summary" in wb.sheetnames: del wb["Summary"]
        ws = wb.create_sheet("Summary")

        qc_columns = [col for col in df.columns if "_OK" in col]
        summary_data = []
        for col in qc_columns:
            total = len(df)
            passed = df[col].sum() if df[col].dtype==bool else sum(df[col]=="True")
            summary_data.append([col, total, passed, total - passed])

        summary_df = pd.DataFrame(summary_data, columns=["Check", "Total", "Passed", "Failed"])
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws.append(r)
        wb.save(output_path)

