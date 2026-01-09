import pandas as pd
import re
import math
import datetime
import os
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

DATE_FORMAT = "%Y-%m-%d"

# Excel color styles
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# ----------------------------- Helpers -----------------------------
def _find_column(df, candidates):
    if df is None:
        return None
    if not isinstance(candidates, list):
        candidates = [candidates]
    cols_lower = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        k = str(cand).lower().strip()
        if k in cols_lower:
            return cols_lower[k]
    return None

def _is_present(val):
    if val is None:
        return False
    try:
        if pd.isna(val):
            return False
    except Exception:
        pass
    if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
        return True
    s = str(val).strip()
    if s == "":
        return False
    if s.lower() in ("nan", "none", "-"):
        return False
    return True

def parse_duration_to_minutes(duration_series):
    results = []
    for item in duration_series:
        if pd.isna(item):
            results.append(np.nan)
            continue
        if isinstance(item, (int, float)):
            results.append(float(item))
            continue
        s = str(item).strip()
        try:
            num = float(s)
            results.append(num)
            continue
        except Exception:
            pass
        parts = s.split(':')
        if len(parts) >= 2:
            try:
                hours = float(re.sub(r"[^0-9.]", "", parts[0])) if parts[0] else 0.0
                minutes = float(re.sub(r"[^0-9.]", "", parts[1])) if parts[1] else 0.0
                seconds = 0.0
                if len(parts) >= 3:
                    seconds = float(re.sub(r"[^0-9.]", "", parts[2])) if parts[2] else 0.0
                total_minutes = (hours * 60) + minutes + (seconds / 60)
                results.append(total_minutes)
            except Exception:
                results.append(np.nan)
        else:
            results.append(np.nan)
    return pd.Series(results, index=duration_series.index)

def to_time_str(val):
    """Convert Excel time/float/time/string to HH:MM:SS string."""
    if pd.isna(val):
        return None

    # if already datetime.time
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")

    # Excel float time (0.0–1.0)
    try:
        if isinstance(val, float) or isinstance(val, int):
            total_seconds = int(val * 24 * 3600)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return f"{h:02}:{m:02}:{s:02}"
    except:
        pass

    # fallback: string
    try:
        t = pd.to_datetime(str(val), errors="coerce")
        if isinstance(t, pd.Timestamp):
            return t.strftime("%H:%M:%S")
    except:
        return None

    return None


def to_date_str(val):
    """Convert Excel date/datetime/string to YYYY-MM-DD."""
    if pd.isna(val):
        return None

    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")

    try:
        d = pd.to_datetime(val, errors="coerce")
        if isinstance(d, pd.Timestamp):
            return d.strftime("%Y-%m-%d")
    except:
        return None

    return None


def combine_parse(date_val, time_val):
    d = to_date_str(date_val)
    t = to_time_str(time_val)
    if not d or not t:
        return pd.NaT
    return pd.to_datetime(f"{d} {t}", errors="coerce")



# ----------------------------- 1️⃣ Detect Monitoring Period -----------------------------
def detect_period_from_rosco(rosco_path):
    # Load the Rosco file
    df = pd.read_excel(rosco_path, header=None)
    
    # 1. Look for the label "Monitoring Periods" in Column B (index 1)
    label_col = df.iloc[:, 1].astype(str)
    period_row_mask = label_col.str.contains("Monitoring Periods", na=False)
    
    if not period_row_mask.any():
        raise ValueError("missing monitoring period label in Column B of Rosco")
    
    # Get the row index (e.g., if found on row 3, index is 2)
    row_idx = period_row_mask.idxmax()
    
    # 2. Check if Column C (index 2) even exists in the loaded dataframe
    # If the user hasn't typed anything in Column C, pandas might not even create the column.
    if df.shape[1] <= 2:
         raise ValueError(f"Missing monitoring period, Please fill the monitoring period in cell C{row_idx + 1} of Rosco")

    # 3. Extract the text from Column C (index 2)
    user_input_text = str(df.iloc[row_idx, 2]).strip()
    
    # Check if the cell is empty or 'nan'
    if not user_input_text or user_input_text.lower() == 'nan':
        raise ValueError(f"missing monitoring period in cell C{row_idx + 1} of Rosco")
    
    # 4. Parse dates from the text (looking for YYYY-MM-DD)
    found = re.findall(r"\d{4}-\d{2}-\d{2}", user_input_text)
    
    if len(found) >= 2:
        start_date = pd.to_datetime(found[0], format=DATE_FORMAT)
        end_date = pd.to_datetime(found[1], format=DATE_FORMAT)
        return start_date, end_date
    else:
        raise ValueError(f"Invalid date format in cell C{row_idx + 1}. Expected two dates (YYYY-MM-DD).")


# ----------------------------- 2️⃣ Load BSR -----------------------------
def detect_header_row_in_sheet(bsr_path, sheet_name):
    df_sample = pd.read_excel(
        bsr_path,
        sheet_name=sheet_name,   #  LOCKED to this sheet
        header=None,
        nrows=200
    )

    for i, row in df_sample.iterrows():
        row_str = " ".join(row.dropna().astype(str)).lower()

        if "region" in row_str and "market" in row_str and "broadcaster" in row_str:
            return i

        if "date" in row_str and ("utc" in row_str or "gmt" in row_str):
            return i

    raise ValueError(
        f"Header not found in sheet '{sheet_name}'"
    )


# def load_bsr(bsr_path):
#     header_row = detect_header_row(bsr_path)
#     df = pd.read_excel(bsr_path, header=header_row)
#     df.columns = [str(c).strip() for c in df.columns]
#     return df

def load_bsr(bsr_path):
    xl = pd.ExcelFile(bsr_path)

    allowed_sheets = {"worksheet", "database"}
    target_sheet = None

    #  Find ONLY worksheet / database
    for sheet in xl.sheet_names:
        if sheet.strip().lower() in allowed_sheets:
            target_sheet = sheet
            break

    if not target_sheet:
        raise ValueError(
            f"No valid sheet ('Worksheet' or 'Database') found in {os.path.basename(bsr_path)}"
        )

    #  Header detection ONLY on the chosen sheet
    header_row = detect_header_row_in_sheet(bsr_path, target_sheet)

    #  Load ONLY that sheet
    df = pd.read_excel(
        bsr_path,
        sheet_name=target_sheet,
        header=header_row
    )

    df.columns = [str(c).strip() for c in df.columns]
    return df

# ----------------------------- 3️⃣ Period Check -----------------------------
def period_check(df, start_date, end_date):
    date_col = next((c for c in df.columns if "date" in str(c).lower()), None)
    if not date_col:
        df["Within_Period_OK"] = True
        df["Within_Period_Remark"] = ""
        return df
    df["Date_checked"] = pd.to_datetime(df[date_col], errors="coerce").dt.date
    df["Within_Period_OK"] = df["Date_checked"].between(start_date.date(), end_date.date())
    df["Within_Period_Remark"] = df["Within_Period_OK"].apply(lambda x: "" if x else "Date outside monitoring period")
    return df


# ----------------------------- 4️⃣ Completeness Check -----------------------------
def completeness_check(df, bsr_cols, rules):
    colmap = {
        "tv_channel": _find_column(df, bsr_cols['tv_channel']),
        "channel_id": _find_column(df, bsr_cols.get('channel_id')),
        "type_of_program": _find_column(df, bsr_cols.get('type_of_program')),
        "match_day": _find_column(df, bsr_cols.get('matchday') or bsr_cols.get('match_day', [])),
        "home_team": _find_column(df, bsr_cols.get('home_team')),
        "away_team": _find_column(df, bsr_cols.get('away_team')),
        "aud_estimates": _find_column(df, bsr_cols.get('aud_estimates')),
        "aud_metered": _find_column(df, bsr_cols.get('aud_metered')),
        "source": _find_column(df, bsr_cols.get('source'))
    }
    df["Completeness_OK"] = True
    df["Completeness_Remark"] = ""
    live_types = set(rules.get('live_types', ['live', 'repeat', 'delayed']))
    relaxed_types = set(rules.get('relaxed_types', ['highlights']))
    for idx, row in df.iterrows():
        missing = []
        for logical, display in [("tv_channel", "TV Channel"), ("channel_id", "Channel ID"),
                                 ("match_day", "Match Day"), ("source", "Source")]:
            colname = colmap.get(logical)
            if colname is None:
                missing.append(f"{display} (column not found)")
            elif not _is_present(row.get(colname)):
                missing.append(display)
        aud_est_col = colmap.get("aud_estimates")
        aud_met_col = colmap.get("aud_metered")
        if not aud_est_col and not aud_met_col:
            missing.append("Audience (Estimates/Metered) (columns not found)")
        else:
            est_present = _is_present(row.get(aud_est_col)) if aud_est_col else False
            met_present = _is_present(row.get(aud_met_col)) if aud_met_col else False
            if not est_present and not met_present:
                missing.append("Both Audience fields are empty")
            elif est_present and met_present:
                missing.append("Both Audience fields are filled")
        type_col = colmap.get("type_of_program")
        prog_type = str(row.get(type_col) or "").strip().lower() if type_col else ""
        home_col, away_col = colmap.get("home_team"), colmap.get("away_team")
        if prog_type in live_types:
            if not home_col: missing.append("Home Team (column not found)")
            elif not _is_present(row.get(home_col)): missing.append("Home Team")
            if not away_col: missing.append("Away Team (column not found)")
            elif not _is_present(row.get(away_col)): missing.append("Away Team")
        elif prog_type not in relaxed_types:
            if home_col and not _is_present(row.get(home_col)): missing.append("Home Team")
            if away_col and not _is_present(row.get(away_col)): missing.append("Away Team")
        if missing:
            df.at[idx, "Completeness_OK"] = False
            df.at[idx, "Completeness_Remark"] = "; ".join(missing)
        else:
            df.at[idx, "Completeness_Remark"] = "All key fields present"
    return df


# ----------------------------- 5️⃣ Overlap / Duplicate / Day Break -----------------------------
def overlap_duplicate_daybreak_check(df, bsr_cols, rules):
    df = df.copy()

    col_channel       = _find_column(df, bsr_cols.get('tv_channel'))
    col_channel_id    = _find_column(df, bsr_cols.get('channel_id'))
    col_market        = _find_column(df, bsr_cols.get('market'))
    col_broadcaster   = _find_column(df, bsr_cols.get('broadcaster'))
    col_start         = _find_column(df, bsr_cols.get('start_time'))
    col_end           = _find_column(df, bsr_cols.get('end_time'))
    # Important: Always use the REAL date column for UTC times
    col_date = _find_column(df, ["Date (UTC/GMT)"])
    if col_date is None:
        col_date = _find_column(df, ["Date"])

    # Minimum columns required
    minimal_required = [col_market, col_date, col_start, col_end]

    if not col_channel and not col_channel_id:
        df["Overlap_OK"] = False
        df["Overlap_Remark"] = "Missing tv_channel and channel_id - cannot run overlap"
        df["Duplicate_OK"] = False
        df["Duplicate_Remark"] = "Missing tv_channel and channel_id - cannot run duplicate"
        df["Daybreak_OK"] = False
        df["Daybreak_Remark"] = "Missing tv_channel and channel_id - cannot run daybreak"
        return df

    if any(c is None for c in minimal_required):
        df["Overlap_OK"] = False
        df["Overlap_Remark"] = "Missing date/start/end - cannot run overlap"
        df["Duplicate_OK"] = False
        df["Duplicate_Remark"] = "Missing date/start/end - cannot run duplicate"
        df["Daybreak_OK"] = False
        df["Daybreak_Remark"] = "Missing date/start/end - cannot run daybreak"
        return df

    compare_channel = col_channel if col_channel else col_channel_id

    df["_start_dt"] = [
        combine_parse(df.at[i, col_date], df.at[i, col_start])
        for i in df.index
    ]
    df["_end_dt"] = [
        combine_parse(df.at[i, col_date], df.at[i, col_end])
        for i in df.index
    ]

    df["_orig_idx"] = df.index

    # Sorting
    sort_by = [compare_channel, col_market, col_date, "_start_dt"]
    df = df.sort_values(by=sort_by, na_position="last").reset_index(drop=True)
    n = len(df)

    # Prepare outputs
    overlap_ok = [pd.NA] * n
    overlap_r  = [""] * n
    duplicate_ok = [True] * n
    duplicate_r  = [""] * n
    daybreak_ok = [pd.NA] * n
    daybreak_r  = [""] * n

    # Duplicate check
    dup_columns = [compare_channel, col_market, col_date, col_start, col_end]
    if col_broadcaster:
        dup_columns.insert(2, col_broadcaster)

    try:
        dup_mask = df.duplicated(subset=dup_columns, keep=False)
    except:
        dup_mask = pd.Series([False] * n)

    for i in range(n):
        if dup_mask.iloc[i]:
            duplicate_ok[i] = False
            duplicate_r[i] = "In-market duplicate (same channel/market/date/start/end)"

    # -------------------------------
    # FIXED Overlap Logic
    # -------------------------------
    df["_grp_channel"] = df[compare_channel].astype(str).str.lower().str.strip()
    df["_grp_market"]  = df[col_market].astype(str).str.lower().str.strip()
    df["_grp_date"]    = df[col_date].astype(str).str.strip()

    grouped = df.groupby(["_grp_channel", "_grp_market", "_grp_date"])

    # -----------------------------
    # Corrected Overlap Check
    # -----------------------------
    for _, grp_idx in grouped.groups.items():
        if len(grp_idx) == 0:
            continue

        # Track PREVIOUS ROW's end time only
        prev_end = None  

        for i in grp_idx:
            curr_start = df.at[i, "_start_dt"]
            curr_end   = df.at[i, "_end_dt"]

            # Missing timestamps → Not Applicable
            if pd.isna(curr_start) or pd.isna(curr_end):
                overlap_ok[i] = pd.NA
                overlap_r[i] = "Not Applicable – missing timestamps"
                continue

            # First event in the group → always OK
            if prev_end is None:
                overlap_ok[i] = True
                overlap_r[i] = "OK (first event in group)"
                prev_end = curr_end              # <-- store this row's end for next comparison
                continue

            # Back-to-back: start equals previous end → OK
            if curr_start == prev_end:
                overlap_ok[i] = True
                overlap_r[i] = "OK – back-to-back scheduling"
                prev_end = curr_end
                continue

            # True overlap: start < previous end
            if curr_start < prev_end:
                overlap_ok[i] = False
                overlap_r[i] = f"Overlap: starts {curr_start.time()} before previous ends {prev_end.time()}"
                # DO NOT update prev_end here (keep the earlier ending to detect longer overlaps)
                continue

            # No overlap → OK
            overlap_ok[i] = True
            overlap_r[i] = "OK"
            prev_end = curr_end

    # -------------------------------
    # Daybreak remains unchanged
    # -------------------------------
    gap_tolerance = rules.get("daybreak_gap_tolerance_min", 5)

    for i in range(1, n):
        prev = df.iloc[i-1]
        curr = df.iloc[i]

        if not (
            str(prev.get(compare_channel)) == str(curr.get(compare_channel))
            and str(prev.get(col_market)) == str(curr.get(col_market))
        ):
            continue

        if pd.isna(prev["_end_dt"]) or pd.isna(curr["_start_dt"]):
            daybreak_ok[i] = pd.NA
            daybreak_r[i] = "Not Applicable – missing timestamps"
            continue

        if prev["_end_dt"].hour >= 23 and curr["_start_dt"].hour <= 1:
            gap = (curr["_start_dt"] - prev["_end_dt"]).total_seconds() / 60
            if 0 <= gap <= gap_tolerance:
                daybreak_ok[i] = True
                daybreak_r[i] = "Valid midnight continuation"
            else:
                daybreak_ok[i] = False
                daybreak_r[i] = f"Invalid continuation gap ({gap:.1f} min)"
        else:
            daybreak_ok[i] = pd.NA
            daybreak_r[i] = "Not Applicable"

    df["Duplicate_OK"] = duplicate_ok
    df["Duplicate_Remark"] = duplicate_r
    df["Overlap_OK"] = overlap_ok
    df["Overlap_Remark"] = overlap_r
    df["Daybreak_OK"] = daybreak_ok
    df["Daybreak_Remark"] = daybreak_r

    return df.sort_values("_orig_idx").drop(
        columns=["_start_dt", "_end_dt", "_orig_idx",
                 "_grp_channel", "_grp_market", "_grp_date"],
        errors="ignore"
    )


# ----------------------------- 6️⃣ Program Category Check -----------------------------
def program_category_check(bsr_path, df, col_map, rules, file_rules):
   # --- Fixture sheet detection (support list or string keywords) ---
    xl = pd.ExcelFile(bsr_path)
    fixture_keywords = file_rules.get("fixture_sheet_keyword", "fixture")
    if not isinstance(fixture_keywords, list):
        fixture_keywords = [fixture_keywords]

    fixture_sheet = None
    for s in xl.sheet_names:
        s_lower = str(s).lower()
        if any(str(kw).lower() in s_lower for kw in fixture_keywords if kw):
            fixture_sheet = s
            break

    if not fixture_sheet:
        df["Program_Category_Expected"] = pd.NA
        df["Program_Category_Actual"] = ""
        df["Program_Category_OK"] = False
        df["Program_Category_Remark"] = "Fixture sheet missing"
        return df

    df_fix = xl.parse(fixture_sheet)

    # ---------- Column detection ----------
    b = col_map["bsr"]
    f = col_map["fixture"]

    col_home_bsr = _find_column(df, b.get("home_team"))
    col_away_bsr = _find_column(df, b.get("away_team"))
    col_date_bsr = _find_column(df, ["Date (UTC/GMT)","Date"])
    col_start_bsr = _find_column(df, ["Start (UTC)","Start"])
    col_progtype = _find_column(df, b.get("type_of_program"))
    col_broadcaster = _find_column(df, b.get("broadcaster"))

    col_combined = _find_column(df, b.get("combined"))
    col_prog_desc = _find_column(df, b.get("program_description"))
    col_prog_title = _find_column(df, b.get("program_title"))
    col_duration = _find_column(df, b.get("duration"))

    col_home_fix = _find_column(df_fix, f.get("home_team"))
    col_away_fix = _find_column(df_fix, f.get("away_team"))
    col_date_fix = _find_column(df_fix, f.get("date"))
    col_start_fix = _find_column(df_fix, f.get("start_time"))

    # ---------- Required columns check ----------
    req = [col_home_bsr, col_away_bsr, col_date_bsr, col_start_bsr]
    if any(c is None for c in req):
        df["Program_Category_Expected"] = pd.NA
        df["Program_Category_Actual"] = df[col_progtype] if col_progtype else ""
        df["Program_Category_OK"] = False
        df["Program_Category_Remark"] = "Missing required columns for LIVE check"
        return df

    # ---------- Helpers ----------
    def clean(x):
        if pd.isna(x):
            return ""
        x = str(x).strip().lower()
        x = x.replace("\u00A0", " ").replace("\u200b", "").strip()
        x = re.sub(r"[^\w\s&]", " ", x)  # keep & since "Magazine & Support" uses it
        return re.sub(r"\s+", " ", x).strip()

    def parse_datetime_candidate(date_raw, time_raw):
        """
        Robust single-row parse producing timezone-naive pd.Timestamp or pd.NaT.
        Tries multiple strategies but does NOT change business logic.
        """
        # quick NA guard
        if (pd.isna(date_raw) or str(date_raw).strip() == "") and (pd.isna(time_raw) or str(time_raw).strip() == ""):
            return pd.NaT

        # Clean text pieces
        d_raw = date_raw
        t_raw = time_raw

        # normalize strings and remove invisible chars
        def norm_obj(o):
            if pd.isna(o):
                return ""
            s = str(o).strip().replace("\u00A0", " ").replace("\u200b", "").strip()
            return s

        d_s = norm_obj(d_raw)
        t_s = norm_obj(t_raw)

        # If time cell accidentally contains letters (e.g., extra text), try to extract leading time
        if t_s and re.search(r"[A-Za-z]", t_s):
            m = re.match(r"^\s*([0-9]{1,2}[:.\-][0-9]{2}(?:[:.\-][0-9]{2})?)", t_s)
            if m:
                t_s = m.group(1)

        # Replace '.' or '-' separators with ':' for times like "20.45" or "20-45"
        t_s = t_s.replace(".", ":").replace("-", ":")

        # If time is an Excel time fraction (0 < t < 1) or numeric fraction string
        try:
            if t_s != "":
                t_as_float = float(t_s)
                if 0 <= t_as_float < 1:  # Excel fraction of day
                    seconds = int(t_as_float * 24 * 3600)
                    hh = seconds // 3600
                    mm = (seconds % 3600) // 60
                    ss = seconds % 60
                    t_s = f"{hh:02}:{mm:02}:{ss:02}"
        except Exception:
            pass

        # If time stored as int like 80000 representing HHMMSS
        try:
            if t_s != "" and re.fullmatch(r"\d{3,6}", t_s):
                val = int(t_s)
                if 0 <= val <= 235959:
                    s6 = str(val).zfill(6)
                    t_s = f"{s6[0:2]}:{s6[2:4]}:{s6[4:6]}"
        except Exception:
            pass

        # Attempt 1: use existing combine_parse helper (handles Excel floats, pandas dates, time strings)
        try:
            ts = combine_parse(d_raw, t_raw)
            # combine_parse may return Timestamp or NaT
            if pd.notna(ts):
                # ensure timezone-naive
                try:
                    if hasattr(ts, "tzinfo") and ts.tzinfo is not None:
                        ts = ts.tz_convert(None)
                except Exception:
                    try:
                        ts = ts.tz_localize(None)
                    except Exception:
                        pass
                return ts
        except Exception:
            pass

        # Attempt 2: If we have a cleaned date+time string, try pd.to_datetime (dayfirst False then True)
        cand_strs = []
        if t_s:
            cand_strs.append(f"{d_s} {t_s}")
        cand_strs.append(d_s)
        for cand in cand_strs:
            for dayfirst in (False, True):
                try:
                    dt = pd.to_datetime(cand, errors="coerce", dayfirst=dayfirst)
                    if pd.notna(dt):
                        # ensure timezone-naive
                        try:
                            if dt.tzinfo is not None:
                                dt = dt.tz_convert(None)
                        except Exception:
                            try:
                                dt = dt.tz_localize(None)
                            except Exception:
                                pass
                        return dt
                except Exception:
                    pass

        # Attempt 3: If date_raw looks numeric (excel serial), convert from serial
        try:
            if isinstance(d_raw, (int, float)) and not math.isnan(d_raw):
                dt_try = pd.to_datetime(d_raw, unit="D", origin="1899-12-30", errors="coerce")
                if pd.notna(dt_try):
                    # if we have t_s, add time part
                    if t_s:
                        try:
                            parts = [int(p) for p in t_s.split(":") if p != ""]
                            hours = parts[0] if len(parts) >= 1 else 0
                            minutes = parts[1] if len(parts) >= 2 else 0
                            seconds = parts[2] if len(parts) >= 3 else 0
                            dt_try = dt_try + pd.Timedelta(hours=hours, minutes=minutes, seconds=seconds)
                        except Exception:
                            pass
                    # make tz-naive
                    try:
                        if dt_try.tzinfo is not None:
                            dt_try = dt_try.tz_convert(None)
                    except Exception:
                        try:
                            dt_try = dt_try.tz_localize(None)
                        except Exception:
                            pass
                    return dt_try
        except Exception:
            pass

        # Attempt 4: if time is numeric fraction only and date is parseable by pd.to_datetime separately
        try:
            base_date = pd.to_datetime(d_s, errors="coerce")
            if pd.notna(base_date) and t_s:
                # parse t_s into hours/mins/secs
                m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", t_s)
                if m:
                    hours = int(m.group(1))
                    minutes = int(m.group(2))
                    seconds = int(m.group(3)) if m.group(3) else 0
                    dt_final = base_date + pd.Timedelta(hours=hours, minutes=minutes, seconds=seconds)
                    try:
                        if dt_final.tzinfo is not None:
                            dt_final = dt_final.tz_convert(None)
                    except Exception:
                        try:
                            dt_final = dt_final.tz_localize(None)
                        except Exception:
                            pass
                    return dt_final
        except Exception:
            pass

        # fallback: NaT
        return pd.NaT

    # ---------- Prepare fixture lookup ----------
    df_fix["_home"] = df_fix[col_home_fix].map(clean)
    df_fix["_away"] = df_fix[col_away_fix].map(clean)
    df_fix["_date"] = pd.to_datetime(df_fix[col_date_fix], errors="coerce").dt.date
    df_fix["_start"] = [
        parse_datetime_candidate(df_fix.at[i, col_date_fix], df_fix.at[i, col_start_fix])
        for i in df_fix.index
    ]

    # ---------- Prepare BSR ----------
    df["_home"] = df[col_home_bsr].map(clean)
    df["_away"] = df[col_away_bsr].map(clean)
    df["_event_key"] = df["_home"] + "||" + df["_away"]
    df["_date"] = pd.to_datetime(df[col_date_bsr], errors="coerce").dt.date
    df["_start"] = [
        parse_datetime_candidate(df.at[i, col_date_bsr], df.at[i, col_start_bsr])
        for i in df.index
    ]
    df["_broad"] = df[col_broadcaster].astype(str).str.lower().str.strip() if col_broadcaster else ""

    # normalize actual to lower-case comparable form
    df["Program_Category_Actual"] = (
        df[col_progtype].astype(str).str.lower().str.strip() if col_progtype else ""
    )

    # combined text for keyword searches
    def get_combined_text(row):
        parts = []
        for c in (col_combined, col_prog_desc, col_prog_title):
            if c and pd.notna(row.get(c, "")):
                parts.append(str(row[c]))
        return " ".join(parts).strip()

    df["_combined_text"] = df.apply(get_combined_text, axis=1).astype(str)
    # duration parse - keep simple behavior
    def parse_duration_minutes(x):
        if pd.isna(x):
            return None
        try:
            return int(float(x))
        except Exception:
            s = str(x).strip()
            m = re.match(r"^(\d+):(\d+)", s)
            if m:
                return int(m.group(1))*60 + int(m.group(2))
            return None
    df["_duration_min"] = df[col_duration].apply(parse_duration_minutes) if col_duration else None

    # keywords & bounds
    highlights_keywords = ["hits", "highlights", "post", "review", "overview", "recap", "summary"]
    magazine_keywords = ["pre", "post", "studio", "interview", "analysis", "previo"]
    dur_min_bound, dur_max_bound = rules.get("flag_duration_min", 10), rules.get("flag_duration_max", 50)

    df["Program_Category_Expected"] = pd.NA
    df["Program_Category_Remark"] = ""

    LIVE_TOL = rules.get("live_tolerance_min", 35)

    # ---------- MAIN LOOP ----------
    for idx, row in df.iterrows():
        ev_key = row["_event_key"]
        h = row["_home"]
        a = row["_away"]
        d = row["_date"]
        bsr_start = row["_start"]
        actual = row["Program_Category_Actual"]
        combined_text = row["_combined_text"]
        dur_min = row["_duration_min"] if col_duration else None

        # ---------- 1) HIGHLIGHTS / MAGAZINE & SUPPORT - OVERRIDE (must NOT use fixture) ----------
        dur_ok = True
        if col_duration:
            if dur_min is None:
                dur_ok = False
            else:
                dur_ok = (dur_min_bound <= dur_min <= dur_max_bound)

        # If actual explicitly labels it, respect that first (normalized)
        if isinstance(actual, str) and actual == "highlights":
            df.at[idx, "Program_Category_Expected"] = "highlights"
            df.at[idx, "Program_Category_Remark"] = "Detected as Highlights (Program type)"
            # override everything else
            continue

        if isinstance(actual, str) and actual in ("magazine", "magazine & support", "magazine & support".lower()):
            # set expected exactly to the normalized form used by actual when present
            df.at[idx, "Program_Category_Expected"] = "magazine & support"
            df.at[idx, "Program_Category_Remark"] = "Detected as Magazine & Support (Program type)"
            continue

        # Keyword+duration based detection (also overrides)
        if dur_ok and any(re.search(rf"\b{re.escape(str(kw))}\b", combined_text.lower()) for kw in highlights_keywords):
            df.at[idx, "Program_Category_Expected"] = "highlights"
            df.at[idx, "Program_Category_Remark"] = f"Detected as Highlights (duration {dur_min} min & keyword match)"
            continue

        if dur_ok and any(re.search(rf"\b{re.escape(str(kw))}\b", combined_text.lower()) for kw in magazine_keywords):
            df.at[idx, "Program_Category_Expected"] = "magazine & support"
            df.at[idx, "Program_Category_Remark"] = f"Detected as Magazine & Support (duration {dur_min} min & keyword match)"
            continue

        # ---------- 2) REPEAT LOGIC (runs before fixture/live/delayed) ----------
        if actual == "repeat":
            same_event = df[df["_event_key"] == ev_key]
            earlier = same_event[
                pd.to_datetime(same_event["_start"], errors="coerce") <
                pd.to_datetime(bsr_start, errors="coerce")
            ]
            if not earlier.empty:
                first_time = pd.to_datetime(earlier["_start"], errors="coerce").min()
                diff = (bsr_start - first_time).total_seconds() / 60
                df.at[idx, "Program_Category_Expected"] = "repeat"
                df.at[idx, "Program_Category_Remark"] = f"Repeat (earlier BSR broadcast exists, {diff:.1f} min earlier)"
                continue
            else:
                df.at[idx, "Program_Category_Expected"] = pd.NA
                df.at[idx, "Program_Category_Remark"] = "Repeat flagged but no earlier BSR broadcast found"
                continue

        # ---------- 3) FIXTURE LOOKUP (only used for live/delayed/repeat) ----------
        fixture_rows = df_fix[
            (df_fix["_home"] == h)
            & (df_fix["_away"] == a)
            & (df_fix["_date"] == d)
        ]

        # If no fixture, do NOT set "No matching fixture found" — leave Expected as NA and blank remark
        if fixture_rows.empty:
            # leave Program_Category_Expected as pd.NA (unless set earlier)
            # do not set remark to "No matching fixture"
            continue

        # fixture exists -> evaluate live/delayed/repeat
        fix_start = fixture_rows["_start"].iloc[0]
        if pd.isna(bsr_start) or pd.isna(fix_start):
            df.at[idx, "Program_Category_Expected"] = pd.NA
            df.at[idx, "Program_Category_Remark"] = "Invalid datetime"
            continue

        diff_min = abs((bsr_start - fix_start).total_seconds() / 60)

        # LIVE
        if diff_min <= LIVE_TOL:
            df.at[idx, "Program_Category_Expected"] = "live"
            df.at[idx, "Program_Category_Remark"] = f"Live (within ±{LIVE_TOL} min)"
            continue

        # DELAYED: only if this is the earliest BSR for the event AND occurs after fixture start
        same_event = df[df["_event_key"] == ev_key]
        earliest = pd.to_datetime(same_event["_start"], errors="coerce").min()

        if pd.to_datetime(bsr_start, errors="coerce") == earliest:
            # ensure it's after fixture start
            if (bsr_start - fix_start).total_seconds() > 0:
                df.at[idx, "Program_Category_Expected"] = "delayed"
                remark = f"Delayed (first telecast outside window; diff {diff_min:.1f} min)"
                if actual != "delayed":
                    remark = remark + f"; note: Program_Type actual='{actual}'"
                df.at[idx, "Program_Category_Remark"] = remark
                continue
            else:
                df.at[idx, "Program_Category_Expected"] = pd.NA
                df.at[idx, "Program_Category_Remark"] = "Broadcast recieved before the fixture start"
                continue
        else:
            # not earliest -> repeat
            df.at[idx, "Program_Category_Expected"] = "repeat"
            later_diff = (bsr_start - earliest).total_seconds() / 60
            df.at[idx, "Program_Category_Remark"] = f"Repeat (first telecast was {later_diff:.1f} min earlier)"
            continue
    # ---------- FINAL OK ----------
    df["Program_Category_OK"] = df["Program_Category_Actual"] == df["Program_Category_Expected"]

    # cleanup internal cols
    df.drop(columns=["_home", "_away", "_event_key", "_date", "_start", "_broad", "_combined_text", "_duration_min"],
            errors="ignore", inplace=True)

    return df

# 8️⃣ Event / Matchday / Competition Check
def check_event_matchday_competition_exact_fixture(df_worksheet,df_fixtures, debug_rows=20):
    """
    Validate Worksheet rows against Fixture List using exact match.

    Logic:
    - Use Event column if present, else fallback to Competition
    - Match on:
        Event/Competition + Matchday + Home Team + Away Team
    - If exact match exists in Fixture List → OK
    - Else → fail with remark
    """

    # ---------- helpers ----------
    def norm(x):
        if pd.isna(x):
            return ""
        return str(x).strip().lower()

    def get_col(df, possible_names):
        for c in df.columns:
            if c.strip().lower() in possible_names:
                return c
        return None

    # ---------- resolve column names ----------
    ws_event_col = get_col(df_worksheet, {"event"})
    ws_comp_col = get_col(df_worksheet, {"competition"})
    ws_matchday_col = get_col(df_worksheet, {"matchday", "match day"})
    ws_home_col = get_col(df_worksheet, {"home team", "hometeam", "home"})
    ws_away_col = get_col(df_worksheet, {"away team", "awayteam", "away"})

    fx_event_col = get_col(df_fixtures, {"event"})
    fx_comp_col = get_col(df_fixtures, {"competition"})
    fx_matchday_col = get_col(df_fixtures, {"matchday", "match day"})
    fx_home_col = get_col(df_fixtures, {"home team", "hometeam", "home"})
    fx_away_col = get_col(df_fixtures, {"away team", "awayteam", "away"})

    # ---------- build fixture lookup set ----------
    fixture_keys = set()

    for _, r in df_fixtures.iterrows():
        event_val = norm(r.get(fx_event_col)) or norm(r.get(fx_comp_col))
        key = (
            event_val,
            norm(r.get(fx_matchday_col)),
            norm(r.get(fx_home_col)),
            norm(r.get(fx_away_col))
        )
        fixture_keys.add(key)

    # ---------- prepare output ----------
    df = df_worksheet.copy()
    df["Event_Matchday_Competition_OK"] = False
    df["Event_Matchday_Competition_Remark"] = ""

    # ---------- row-wise validation ----------
    for idx, r in df.iterrows():
        event_val = norm(r.get(ws_event_col)) or norm(r.get(ws_comp_col))
        matchday = norm(r.get(ws_matchday_col))
        home = norm(r.get(ws_home_col))
        away = norm(r.get(ws_away_col))

        key = (event_val, matchday, home, away)

        if key in fixture_keys and all(key):
            df.at[idx, "Event_Matchday_Competition_OK"] = True
            df.at[idx, "Event_Matchday_Competition_Remark"] = "OK"
        else:
            df.at[idx, "Event_Matchday_Competition_OK"] = False
            df.at[idx, "Event_Matchday_Competition_Remark"] = "Exact match not found in fixture"

    # ---------- debug ----------
    print("=== Exact Fixture Match QC (sample rows) ===")
    for i in range(min(debug_rows, len(df))):
        r = df.iloc[i]
        print(
            f"[Row {i}] Event/Comp='{norm(r.get(ws_event_col)) or norm(r.get(ws_comp_col))}' | "
            f"MD='{r.get(ws_matchday_col)}' | "
            f"Home='{r.get(ws_home_col)}' | Away='{r.get(ws_away_col)}' | "
            f"OK={r['Event_Matchday_Competition_OK']} | "
            f"Remark={r['Event_Matchday_Competition_Remark']}"
        )
    print("=== End QC ===\n")

    return df

# -----------------------------------------------------------
# 9️⃣ Market / Channel / Program / Duration Consistency Check

def market_channel_consistency_check(df_bsr, rosco_path, col_map, file_rules):
    logging.info("🔍 Starting Market & Channel Consistency Check...")
    bsr_cols = col_map['bsr']
    rosco_cols = col_map.get('rosco', {})
    def normalize_channel(name):
        if pd.isna(name) or name is None:
            return ""
        s = str(name)
        s = re.sub(r"\(.*?\)|\[.*?\]", "", s)
        s = re.split(r"[-–—]", s)[0]
        s = re.sub(r"[^0-9a-zA-Z\s]", " ", s)
        return re.sub(r"\s+", " ", s).strip().lower()
    rosco_df = None
    if rosco_path:
        try:
            xls = pd.ExcelFile(rosco_path)
            ignore_sheet = file_rules.get('rosco_ignore_sheet', 'general')
            sheet_name = next((s for s in xls.sheet_names if ignore_sheet not in s.lower()), None)
            if sheet_name:
                rosco_df = xls.parse(sheet_name)
            else:
                logging.warning(f"No valid sheet found in ROSCO (ignoring '{ignore_sheet}').")
        except Exception as e:
            logging.error(f"Error loading ROSCO file: {e}")
            df_bsr["Market_Channel_Consistency_OK"] = False
            df_bsr["Market_Channel_Program_Remark"] = f"Error loading ROSCO: {e}"
            return df_bsr
    valid_pairs = set()
    rosco_country_col = rosco_cols.get('channel_country', 'ChannelCountry')
    rosco_name_col = rosco_cols.get('channel_name', 'ChannelName')
    if rosco_df is not None and not rosco_df.empty and {rosco_country_col, rosco_name_col}.issubset(rosco_df.columns):
        for _, row in rosco_df.iterrows():
            market = str(row[rosco_country_col]).strip().lower()
            channel = normalize_channel(row[rosco_name_col])
            if market and channel:
                valid_pairs.add((market, channel))
        logging.info(f"Loaded {len(valid_pairs)} valid Market+Channel pairs from ROSCO.")
    df_bsr["Market_Channel_Consistency_OK"] = True
    df_bsr["Market_Channel_Program_Remark"] = "OK"
    bsr_market_col = _find_column(df_bsr, bsr_cols.get('market'))
    bsr_channel_col = _find_column(df_bsr, bsr_cols.get('tv_channel'))
    if not bsr_market_col or not bsr_channel_col:
        logging.error("Market/Channel Check: BSR columns not found. Skipping.")
        df_bsr["Market_Channel_Consistency_OK"] = False
        df_bsr["Market_Channel_Program_Remark"] = "BSR columns not found"
        return df_bsr
    for idx, row in df_bsr.iterrows():
        remarks = []
        market = str(row.get(bsr_market_col, "")).strip().lower()
        channel = str(row.get(bsr_channel_col, "")).strip()
        if not market or not channel:
            df_bsr.at[idx, "Market_Channel_Consistency_OK"] = False
            remarks.append("Missing market or channel")
        elif valid_pairs:
            if (market, normalize_channel(channel)) not in valid_pairs:
                df_bsr.at[idx, "Market_Channel_Consistency_OK"] = False
                remarks.append("Market+Channel not found in ROSCO")
        df_bsr.at[idx, "Market_Channel_Program_Remark"] = "; ".join(remarks) if remarks else "OK"
    logging.info("✅ Market & Channel Consistency Check completed.")
    return df_bsr

# -----------------------------------------------------------
# 10️⃣ Domestic Market Coverage Check
def domestic_market_check(df_worksheet, bsr_cols, monitoring_start_date=None, debug=False):
    df = df_worksheet.copy()
    df["Domestic_Market_Coverage_OK"] = True
    df["Domestic_Market_Remark"] = ""
    col_comp = _find_column(df, bsr_cols.get('competition', ['Competition']))
    col_mkt = _find_column(df, bsr_cols.get('market', ['Market']))
    col_date = _find_column(df, bsr_cols.get('date', ['Date']))
    col_prog_type = _find_column(df, bsr_cols.get('type_of_program', ['Type of Program']))
    if not all([col_comp, col_mkt, col_date, col_prog_type]):
        df["Domestic_Market_Coverage_OK"] = False
        df["Domestic_Market_Remark"] = "Skipped: Missing core BSR columns in file/config."
        return df
    DOMESTIC_MAP = {
        "premier league": ["united kingdom", "england"],
        "epl": ["united kingdom", "england"],
        "la liga": ["spain"],
        "bundesliga": ["germany", "deutschland"],
        "serie a": ["italy"],
        "ligue 1": ["france"]
    }
    monitoring_start = None
    if monitoring_start_date is not None:
        try:
            monitoring_start = pd.to_datetime(monitoring_start_date).date()
        except Exception:
            monitoring_start = None
    for idx, row in df.iterrows():
        comp = str(row.get(col_comp, "")).strip().lower()
        market = str(row.get(col_mkt, "")).strip().lower()
        date_raw = row.get(col_date)
        try:
            row_date = pd.to_datetime(date_raw).date()
        except Exception:
            row_date = None
        if monitoring_start and row_date and row_date < monitoring_start:
            continue
        domestic_markets = []
        for comp_kw, markets in DOMESTIC_MAP.items():
            if comp_kw in comp:
                domestic_markets = markets
                break
        if not domestic_markets:
            continue
        market_ok = any(dm in market for dm in domestic_markets)
        if not market_ok:
            df.at[idx, "Domestic_Market_Coverage_OK"] = False
            df.at[idx, "Domestic_Market_Remark"] = f"Missing domestic coverage. Expected one of: {domestic_markets}"
        else:
            df.at[idx, "Domestic_Market_Remark"] = "OK"
    return df

# -----------------------------------------------------------
# 11️⃣ Rates & Ratings Check
# --------------------------------------------
def rates_and_ratings_check(df, bsr_cols):
    est_col = _find_column(df, bsr_cols.get('aud_estimates'))
    met_col = _find_column(df, bsr_cols.get('aud_metered'))
    est_col_exists = est_col is not None and est_col in df.columns
    met_col_exists = met_col is not None and met_col in df.columns
    if est_col is None:
        est_col = "Audience_Estimates_Dummy"
        df[est_col] = np.nan
        logging.warning("Rates/Ratings Check: Audience Estimates column not found.")
    if met_col is None:
        met_col = "Audience_Metered_Dummy"
        df[met_col] = np.nan
        logging.warning("Rates/Ratings Check: Audience Metered column not found.")
    present_est = df[est_col].apply(_is_present)
    present_met = df[met_col].apply(_is_present)
    both_empty_mask = (~present_est) & (~present_met)
    both_present_mask = (present_est) & (present_met)
    exactly_one_mask = (present_est ^ present_met)
    df["Rates_Ratings_QC_OK"] = True
    df["Rates_Ratings_QC_Remark"] = ""
    df.loc[both_empty_mask, "Rates_Ratings_QC_OK"] = False
    df.loc[both_empty_mask, "Rates_Ratings_QC_Remark"] = "Missing audience ratings (both empty)"
    df.loc[both_present_mask, "Rates_Ratings_QC_OK"] = False
    df.loc[both_present_mask, "Rates_Ratings_QC_Remark"] = "Invalid: both metered and estimated present"
    df.loc[exactly_one_mask, "Rates_Ratings_QC_OK"] = True
    df.loc[exactly_one_mask, "Rates_Ratings_QC_Remark"] = "Valid: one rating source available"
    if est_col == "Audience_Estimates_Dummy" and est_col in df.columns:
        df.drop(columns=[est_col], inplace=True)
    if met_col == "Audience_Metered_Dummy" and met_col in df.columns:
        df.drop(columns=[met_col], inplace=True)
    return df

# -----------------------------------------------------------
# 12️⃣ Comparison of Duplicated Markets
def duplicated_market_check(df_bsr, macro_path, project, col_map, file_rules, debug=False):

    result_col = "Duplicated_Markets_Check_OK"
    remark_col = "Duplicated_Markets_Remark"

    df_bsr[result_col] = pd.NA
    df_bsr[remark_col] = "Not Applicable"

    league_keyword = str(project.get("league_keyword", "F24 Spain")).lower()
    bsr_cols = col_map["bsr"]
    macro_cols = col_map["macro"]

    if not macro_path or not os.path.exists(macro_path):
        df_bsr[result_col] = False
        df_bsr[remark_col] = "Macro file missing"
        return df_bsr


    # -------------------------------------------------------
    # 🔥 STEP 1 — Load Excel WITHOUT trusting header_row
    # -------------------------------------------------------
    try:
        xl = pd.ExcelFile(macro_path, engine="openpyxl")

        # Pick correct sheet
        preferred = file_rules.get("macro_sheet_name", "Data Core").lower()
        sheet = next((s for s in xl.sheet_names if s.lower() == preferred), xl.sheet_names[0])

        # Read top 20 rows without header
        tmp = pd.read_excel(macro_path, sheet_name=sheet, header=None, nrows=20, dtype=str)

        required_cols = ["Projects", "Orig Market", "Orig Channel", "Dup Market", "Dup Channel"]

        header_row_index = None

        # 🔍 Find the row where all required column names appear
        for i in range(len(tmp)):
            row_vals = [str(x).strip().lower() for x in list(tmp.iloc[i].values)]
            if all(any(req.lower() == val for val in row_vals) for req in required_cols):
                header_row_index = i
                break

        if header_row_index is None:
            df_bsr[result_col] = False
            df_bsr[remark_col] = "Could not locate header row in macro file."
            return df_bsr

        # Now correctly load macro_df using detected header row
        macro_df = pd.read_excel(
            macro_path,
            sheet_name=sheet,
            header=header_row_index,
            dtype=str,
            engine="openpyxl"
        )

        macro_df.columns = [str(c).strip() for c in macro_df.columns]

    except Exception as e:
        df_bsr[result_col] = False
        df_bsr[remark_col] = f"Macro load error: {e}"
        return df_bsr


    # -------------------------------------------------------
    # 🔥 STEP 2 — Find required columns reliably
    # -------------------------------------------------------
    def find_col(df, key):
        if isinstance(key, list):
            candidates = key
        else:
            candidates = [key]

        lower = {c.lower(): c for c in df.columns}
        for cand in candidates:
            c = str(cand).strip().lower()
            if c in lower:
                return lower[c]
        return None

    proj_col = find_col(macro_df, macro_cols["projects"])
    orig_mkt_col = find_col(macro_df, macro_cols["orig_market"])
    orig_ch_col = find_col(macro_df, macro_cols["orig_channel"])
    dup_mkt_col = find_col(macro_df, macro_cols["dup_market"])
    dup_ch_col = find_col(macro_df, macro_cols["dup_channel"])

    missing = [col for col in [proj_col, orig_mkt_col, orig_ch_col, dup_mkt_col, dup_ch_col] if col is None]
    if missing:
        df_bsr[result_col] = False
        df_bsr[remark_col] = "Macro file columns not found (after auto-detect)."
        return df_bsr


    # -------------------------------------------------------
    # 🔥 STEP 3 — Filter by project keyword
    # -------------------------------------------------------
    macro_df = macro_df[
        macro_df[proj_col].astype(str).str.lower().str.contains(league_keyword, na=False)
    ]

    if macro_df.empty:
        df_bsr[result_col] = pd.NA
        df_bsr[remark_col] = f"No duplication rules found for {league_keyword}"
        return df_bsr


    # -------------------------------------------------------
    # 🔥 STEP 4 — Run duplication checks (unchanged logic)
    # -------------------------------------------------------
    mkt_col = find_col(df_bsr, bsr_cols["market"])
    ch_col = find_col(df_bsr, bsr_cols["tv_channel"])
    comp_col = find_col(df_bsr, bsr_cols["competition"])
    evt_col = find_col(df_bsr, bsr_cols["event"])

    in_league = (
        df_bsr[comp_col].astype(str).str.lower().str.contains(league_keyword, na=False)
        | df_bsr[evt_col].astype(str).str.lower().str.contains(league_keyword, na=False)
    )

    df_bsr.loc[~in_league, result_col] = pd.NA
    df_bsr.loc[~in_league, remark_col] = "Not Applicable"

    df_league = df_bsr[in_league].copy()

    for _, r in macro_df.iterrows():
        orig_market = str(r[orig_mkt_col]).strip().lower()
        orig_channel = str(r[orig_ch_col]).strip().lower()
        dup_market = str(r[dup_mkt_col]).strip().lower()
        dup_channel = str(r[dup_ch_col]).strip().lower()

        orig_rows = df_league[
            (df_league[mkt_col].str.lower() == orig_market) &
            (df_league[ch_col].str.lower() == orig_channel)
        ]
        dup_rows = df_league[
            (df_league[mkt_col].str.lower() == dup_market) &
            (df_league[ch_col].str.lower() == dup_channel)
        ]

        orig_events = set(orig_rows[evt_col].dropna().str.lower().str.strip())
        dup_events = set(dup_rows[evt_col].dropna().str.lower().str.strip())

        if not orig_events:
            status = pd.NA
            remark = f"No events found for {orig_market}/{orig_channel}"
        elif orig_events.issubset(dup_events):
            status = True
            remark = f"All {len(orig_events)} events duplicated"
        else:
            missing = orig_events - dup_events
            status = False
            remark = f"Missing {len(missing)} events"

        mask = (
            (df_bsr[mkt_col].str.lower() == orig_market) &
            (df_bsr[ch_col].str.lower() == orig_channel) &
            in_league
        ) | (
            (df_bsr[mkt_col].str.lower() == dup_market) &
            (df_bsr[ch_col].str.lower() == dup_channel) &
            in_league
        )

        df_bsr.loc[mask, result_col] = status
        df_bsr.loc[mask, remark_col] = remark

    return df_bsr
# -----------------------------------------------------------
# 13️⃣ Country & Channel IDs Check
def country_channel_id_check(df, bsr_cols):
    """
    Check consistency of channel IDs per (market, tv_channel) pair.

    RULE:
    - For each (Market, TV-Channel) pair → must have exactly ONE unique non-blank Channel ID.
    - If same pair appears with different non-blank Channel IDs → inconsistent.
    - If the only channel_id is blank → inconsistent (Missing channel ID).
    - Same TV-Channel across different markets is allowed (treated independently).

    Adds these columns:
        Market_Channel_ID_OK (bool)
        Market_Channel_ID_Remark (str)
    """

    df = df.copy()  # work on a copy to avoid side-effects
    df["Market_Channel_ID_OK"] = True
    df["Market_Channel_ID_Remark"] = "OK"

    # Resolve column names (use _find_column)
    ch_col = _find_column(df, bsr_cols.get("tv_channel"))
    ch_id_col = _find_column(df, bsr_cols.get("channel_id"))
    mkt_col = _find_column(df, bsr_cols.get("market"))

    if not all([ch_col, ch_id_col, mkt_col]):
        logging.warning("Country/Channel ID Check: Missing required columns. Skipping.")
        df["Market_Channel_ID_OK"] = False
        df["Market_Channel_ID_Remark"] = "Check skipped: missing required columns"
        return df

    def norm(x):
        if pd.isna(x) or x is None:
            return ""
        return str(x).strip()

    # Build mapping: (market, tv_channel) → set(channel_ids) & row indices
    pair_ids = {}
    pair_idxs = {}

    for idx, row in df.iterrows():
        market = norm(row.get(mkt_col, ""))
        channel = norm(row.get(ch_col, ""))
        channel_id = norm(row.get(ch_id_col, ""))

        # Normalize for comparisons (lower-case for market and channel)
        key = (market.lower(), channel.lower())

        pair_ids.setdefault(key, set()).add(channel_id)
        pair_idxs.setdefault(key, []).append(idx)

    # Evaluate each pair
    for key, id_set in pair_ids.items():
        idxs = pair_idxs.get(key, [])
        # consider only non-blank IDs for uniqueness check
        non_blank_ids = {cid for cid in id_set if cid != ""}

        inconsistent = False
        remark = "OK"

        if len(non_blank_ids) == 0:
            inconsistent = True
            remark = "Missing channel ID"
        elif len(non_blank_ids) > 1:
            inconsistent = True
            # keep blanks visible as <BLANK> if present
            ids_list = [cid if cid != "" else "<BLANK>" for cid in sorted(id_set)]
            # include market/channel for clarity in remark
            market_display, channel_display = key
            remark = f"Conflicting Channel IDs for {channel_display} in market {market_display}: {', '.join(ids_list)}"
        else:
            inconsistent = False
            remark = "OK"

        for i in idxs:
            df.at[i, "Market_Channel_ID_OK"] = not inconsistent
            df.at[i, "Market_Channel_ID_Remark"] = remark

    return df

# -----------------------------------------------------------
# ✅ Excel Coloring for True/False checks
def color_excel(output_path, df):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx+1 for idx, name in enumerate(headers)}

    qc_columns = [col for col in df.columns if col.endswith("_OK")]

    for col_name in qc_columns:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val in [True, "True"]:
                    cell.fill = GREEN_FILL
                elif val in [False, "False"]:
                    cell.fill = RED_FILL

    wb.save(output_path)
# -----------------------------------------------------------
# Summary Sheet
def generate_summary_sheet(output_path, df):
    wb = load_workbook(output_path)
    if "Summary" in wb.sheetnames: del wb["Summary"]
    ws = wb.create_sheet("Summary")

    qc_columns = [col for col in df.columns if "_OK" in col]
    summary_data = []
    for col in qc_columns:
        total = len(df)
        passed = df[col].sum() if df[col].dtype==bool else sum(df[col]=="True")
        summary_data.append([col, total, passed, total - passed])

    summary_df = pd.DataFrame(summary_data, columns=["Check", "Total", "Passed", "Failed"])
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    wb.save(output_path)